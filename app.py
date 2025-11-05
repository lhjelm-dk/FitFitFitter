"""
FitFitFitter - Interactive Distribution Fitting and Statistical Analysis Tool
A Streamlit application for analyzing data and fitting probability distributions.
"""

import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
import os
from pathlib import Path
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.chart import ScatterChart, LineChart, BarChart, Reference, Series
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.axis import DateAxis, NumericAxis
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Set seaborn color palette
sns.set_palette("husl")

# Import utilities
from utils.fitting import (
    rank_distributions, 
    format_params, 
    get_param_names, 
    get_default_bounds,
    DISTRIBUTIONS
)

# Page configuration
st.set_page_config(
    page_title="FitFitFitter - Distribution Fitting Tool",
    layout="wide"
)

# Title
st.title("FitFitFitter")
st.markdown("**Interactive Distribution Fitting and Statistical Analysis Tool**")

# Documentation sections
with st.expander("Quick Start Guide", expanded=False):
    st.markdown("""
    ### Getting Started
    
    1. **Input Your Data**
       - **Upload File**: Supports multiple formats (CSV, TXT, Excel .xlsx/.xls, JSON, Parquet)
         - After uploading, a preview table shows the first 10 rows
         - Select which numeric column to analyze (auto-detected if multiple columns exist)
         - Optional data cleaning: remove non-numeric values, zeros, or clip outliers above P99
         - Your uploaded dataset is cached - use "Reload Last Data" if you refresh the page
       - **Paste Values**: Paste comma-separated values or tabular data (e.g., column from Excel)
       - **Use Test Data**: Explore the tool with the provided example dataset
       - Ensure you have at least 10 valid (non-negative) numeric values
    
    2. **View Descriptive Statistics**
       - Once data is loaded, basic statistics are automatically displayed
       - Includes: count, mean, min, max, standard deviation, and percentiles (P10, P50, P90)
       - **Skewness** and **Kurtosis** describe data shape (asymmetry and tail weight)
       - Use these shape metrics to guide distribution selection
    
    3. **Automatic Distribution Fitting**
       - Click "Fit Distributions" to automatically fit multiple distributions to your data
       - Results are ranked by **Kolmogorov-Smirnov (KS) statistic** (lower = better fit)
       - Review the top fits with their parameters and goodness-of-fit statistics
       - View **Q-Q plots, P-P plots, Histogram+PDF overlays, and CDF Difference plots** for visual diagnostics
       - Get **LaTeX and Excel formulas** for each fitted distribution
       - Use plot options to adjust the number of distributions shown, x-axis range, bins, or log scale
       - 📖 **For detailed interpretation**: See "Understanding Goodness-of-Fit Tests" section below
    
    4. **Interactive Manual Fitting**
       - Select a distribution type from the dropdown
       - Adjust parameters using the sliders
       - Watch the curve update in real-time on the inverse CDF plot
       - Compare your manual fit with the empirical data
    
    ### Understanding the Plots
    
    - **Red dots**: Empirical inverse cumulative distribution (1 - CDF) from your data
    - **Blue line**: Fitted model's inverse CDF curve
    - **Coral bars**: Histogram (density function) of your empirical data - shows the distribution shape
    - **Brown filled area**: Probability density function (PDF) of the fitted distribution - shows the theoretical distribution shape
    - **Log scale**: Useful for datasets with wide value ranges
    
    The plots show both cumulative (1-CDF) and density functions, allowing you to see how well the fitted distribution matches your data in both views.
    
    ### Interpreting Results (Quick Reference)
    
    The tool provides several goodness-of-fit statistics to help you evaluate distribution fits:
    
    - **KS Statistic** (Primary ranking): Lower values = better fit. Ranks distributions automatically.
    - **P-value**: Higher values suggest better fit. Typically p > 0.05 indicates acceptable fit.
    - **AIC**: Lower values = better relative fit (only meaningful when comparing models on same data).
    - **Chi-square Test**: Compares observed vs. expected frequencies. p ≥ 0.05 suggests good fit.
    - **Anderson-Darling**: Available for Normal, Exponential, Logistic. Lower values = better fit.
    - **Visual Diagnostics**: Q-Q and P-P plots show quantile and probability matching. CDF Difference plot shows systematic biases in cumulative probability estimation.
    
    **For detailed explanations, mathematical foundations, and interpretation guidelines**, see the **"Understanding Goodness-of-Fit Tests"** section below.
    """)

with st.expander("Understanding Goodness-of-Fit Tests", expanded=False):
    st.markdown("""
    ### Overview
    
    This section provides detailed explanations of the statistical tests used to evaluate distribution fits. The tool ranks distributions by the **Kolmogorov-Smirnov (KS) statistic**, with additional tests (AIC, Chi-square, Anderson-Darling) providing complementary information for comprehensive assessment.
    
    
    ---
    
    ### Kolmogorov-Smirnov (KS) Test (Primary Ranking Criterion)
    
    #### Mathematical Foundation
    
    The KS statistic measures the **supremum** (maximum) distance between the empirical cumulative distribution function (ECDF) and the theoretical cumulative distribution function (CDF) of the fitted distribution:
    
    $$D_n = \\sup_x |F_n(x) - F(x)|$$
    
    where:
    - $F_n(x)$ is the empirical CDF: $F_n(x) = \\frac{1}{n}\\sum_{i=1}^{n} I_{[x_i \\leq x]}$
    - $F(x)$ is the theoretical CDF of the fitted distribution
    - $n$ is the sample size
    - $I_{[x_i \\leq x]}$ is the indicator function
    
    **Key properties**:
    - **Non-parametric**: Makes no assumptions about the underlying distribution
    - **Distribution-free**: Critical values depend only on sample size, not the hypothesized distribution
    - **Sensitive to all deviations**: Detects differences in location, scale, and shape
    
    #### Interpretation Guidelines
    
    | KS Statistic | Interpretation | Action |
    |:------------|:--------------|:------|
    | < 0.05 | Excellent fit | Distribution is a very good model |
    | 0.05 - 0.10 | Good fit | Distribution adequately describes the data |
    | 0.10 - 0.20 | Acceptable fit | Distribution reasonably matches data, but consider alternatives |
    | > 0.20 | Poor fit | Distribution does not fit well; try other distributions |
    
    **Note**: These thresholds are general guidelines. Consider your specific application context and sample size.
    
    #### KS Statistic vs P-value
    
    - **KS Statistic ($D_n$)**: The actual maximum distance between ECDF and CDF (lower = better fit)
    - **P-value**: The probability of observing a KS statistic this large or larger under the null hypothesis that the data comes from the fitted distribution
      - **Null hypothesis ($H_0$)**: Data follows the fitted distribution
      - **High p-value** (e.g., p > 0.05): Cannot reject $H_0$; data is consistent with the distribution
      - **Low p-value** (e.g., p < 0.05): Reject $H_0$; data is unlikely to come from this distribution
    
    **Ideal outcome**: Low KS statistic + High p-value indicates strong evidence that the distribution fits the data well.
    
    #### Why KS Test is Used for Ranking
    
    1. **Visual correspondence**: The KS statistic directly corresponds to the maximum gap you see between data points and fitted curve on the inverse CDF plot
    2. **Robustness**: Works well across different sample sizes (though critical values adjust with n)
    3. **Comparability**: Allows ranking different distributions on the same scale
    4. **No distributional assumptions**: Works for any continuous distribution
    
    #### Visual Interpretation
    
    On the inverse CDF plot (1 - CDF):
    - **Red dots**: Your empirical data points
    - **Colored lines**: Fitted distribution curves
    - **KS statistic**: The maximum vertical distance between the data points and the fitted curve
    - **Better fits**: Show fitted lines closely following data points with minimal gaps
    
    #### Further Reading
    
    - **Wikipedia**: [Kolmogorov–Smirnov test](https://en.wikipedia.org/wiki/Kolmogorov%E2%80%93Smirnov_test)
    - **Scipy Documentation**: [scipy.stats.kstest](https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.kstest.html)
    - **Statistical Text**: Massey, F. J. (1951). "The Kolmogorov-Smirnov Test for Goodness of Fit." *Journal of the American Statistical Association*, 46(253), 68-78.
    
    ---
    
    ### Akaike Information Criterion (AIC)
    
    #### Mathematical Foundation
    
    AIC is a model selection criterion that balances goodness-of-fit with model complexity:
    
    $$\\text{AIC} = 2k - 2\\ln(L)$$
    
    where:
    - $k$ is the number of parameters in the model
    - $L$ is the maximum likelihood value
    - $\\ln(L)$ is the log-likelihood
    
    **Derivation**: Based on information theory, AIC estimates the relative information loss when using a model to approximate reality. The penalty term ($2k$) prevents overfitting by penalizing models with more parameters.
    
    #### Interpretation
    
    - **Lower AIC = Better relative fit**: Among models fitted to the same data, the model with the lowest AIC is preferred
    - **Relative measure**: AIC values are only meaningful when comparing models on the **same dataset**
    - **Absolute values**: The absolute AIC value has no meaning; only differences matter
    - **Rule of thumb**: A difference of 2-6 suggests moderate evidence for the better model; >10 suggests strong evidence
    
    #### When to Use AIC
    
    - **Model comparison**: Choose between distributions with different numbers of parameters
    - **Complexity penalty**: Prefer simpler models when fit is similar (parsimony principle)
    - **Not for**: Absolute goodness-of-fit assessment (use KS test for that)
    
    #### Limitations
    
    - Requires maximum likelihood estimation (used in this tool)
    - Assumes large sample sizes (asymptotic property)
    - Only valid for comparing models on identical data
    
    #### Further Reading
    
    - **Wikipedia**: [Akaike Information Criterion](https://en.wikipedia.org/wiki/Akaike_information_criterion)
    - **Burnham & Anderson (2002)**: *Model Selection and Multimodel Inference: A Practical Information-Theoretic Approach*
    
    ---
    
    ### Chi-Square Goodness-of-Fit Test
    
    #### Mathematical Foundation
    
    The Chi-square test compares observed frequencies in histogram bins with expected frequencies under the fitted distribution:
    
    $$\\chi^2 = \\sum_{i=1}^{k} \\frac{(O_i - E_i)^2}{E_i}$$
    
    where:
    - $O_i$ is the observed frequency in bin $i$
    - $E_i$ is the expected frequency in bin $i$ under the fitted distribution
    - $k$ is the number of bins
    
    Under the null hypothesis that data follows the fitted distribution, the test statistic follows a Chi-square distribution with degrees of freedom:
    
    $$\\text{df} = k - 1 - p$$
    
    where $p$ is the number of fitted parameters.
    
    #### Interpretation
    
    | P-value | Interpretation | Meaning |
    |:--------|:--------------|:--------|
    | ≥ 0.05 | Good fit | Cannot reject $H_0$; observed frequencies match expected frequencies well |
    | < 0.05 | Poor fit | Reject $H_0$; observed and expected frequencies differ significantly |
    
    **Null hypothesis ($H_0$)**: Data follows the fitted distribution (frequencies match expected values)
    
    #### Limitations and Considerations
    
    1. **Bin dependence**: Results sensitive to bin number and boundaries
    2. **Sample size**: Requires sufficient expected frequencies (typically ≥ 5 per bin)
    3. **Adaptive binning**: This tool uses adaptive binning to ensure valid test conditions
    4. **Discrete approximation**: Treats continuous data as discrete bins, which may lose information
    
    **Best practices**:
    - Most reliable with large samples (n > 50)
    - Use as supplementary test alongside KS test
    - Consider visual inspection of histogram + PDF overlay
    
    #### Further Reading
    
    - **Wikipedia**: [Pearson's Chi-squared test](https://en.wikipedia.org/wiki/Pearson%27s_chi-squared_test)
    - **Statistical Text**: Snedecor, G. W., & Cochran, W. G. (1989). *Statistical Methods* (8th ed.), Chapter 6
    
    ---
    
    ### Anderson-Darling Test
    
    #### Mathematical Foundation
    
    The Anderson-Darling test is a modification of the KS test that gives more weight to the tails of the distribution:
    
    $$A^2 = -n - \\frac{1}{n}\\sum_{i=1}^{n}(2i-1)[\\ln F(x_i) + \\ln(1-F(x_{n+1-i}))]$$
    
    where $F(x_i)$ is the CDF of the fitted distribution evaluated at the $i$-th ordered data point.
    
    **Key difference from KS test**: The weight function $(2i-1)$ emphasizes deviations in the tails, making it more sensitive to extreme values.
    
    #### Interpretation
    
    - **Lower values = Better fit**: Unlike KS, Anderson-Darling has distribution-specific critical values
    - **Tail sensitivity**: More likely to detect deviations in distribution tails than KS test
    - **Use case**: Particularly useful when tail behavior is important (e.g., risk assessment, extreme events)
    
    #### Availability
    
    Currently implemented for:
    - **Normal distribution**: Tests normality
    - **Exponential distribution**: Tests exponentiality
    - **Logistic distribution**: Tests logistic distribution
    
    For other distributions, the test statistic is calculated but critical values may not be available.
    
    #### Further Reading
    
    - **Wikipedia**: [Anderson-Darling test](https://en.wikipedia.org/wiki/Anderson%E2%80%93Darling_test)
    - **Original Paper**: Anderson, T. W., & Darling, D. A. (1954). "A Test of Goodness of Fit." *Journal of the American Statistical Association*, 49(268), 765-769.
    - **Scipy Documentation**: [scipy.stats.anderson](https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.anderson.html)
    
    ---
    
    ### Visual Diagnostic Plots
    
    #### Q-Q Plot (Quantile-Quantile Plot)
    
    Plots empirical quantiles (from your data) against theoretical quantiles (from the fitted distribution).
    
    **Interpretation**:
    - **Points on diagonal line**: Good fit (quantiles match)
    - **Systematic deviations**: Indicates distribution mismatch
    - **Curved patterns**: Suggest distribution shape mismatch (e.g., skewness, tail behavior)
    - **Outliers**: Points far from diagonal indicate extreme values not well captured
    
    #### P-P Plot (Probability-Probability Plot)
    
    Plots empirical cumulative probabilities against theoretical cumulative probabilities.
    
    **Interpretation**:
    - **Points on diagonal line**: Good fit (probabilities match)
    - **S-shaped curve**: Indicates location/scale issues
    - **U-shaped or inverted U**: Suggests shape parameter mismatch
    
    **Q-Q vs P-P**: Q-Q plots are more sensitive to tail deviations; P-P plots are more sensitive to center deviations.
    
    #### Histogram + PDF Overlay
    
    Direct visual comparison of:
    - **Histogram**: Empirical density (what your data looks like)
    - **PDF curve**: Theoretical density (what the distribution predicts)
    
    **Interpretation**: Good fit shows histogram bars aligning with PDF curve shape.
    
    #### CDF Difference Plot
    
    Plots the difference between the theoretical CDF (from fitted distribution) and the empirical CDF (cumulative frequency of sample data) at each data point:
    
    $$\\text{Difference} = F_{\\text{theoretical}}(x_i) - F_{\\text{empirical}}(x_i)$$
    
    where:
    - $F_{\\text{theoretical}}(x_i)$ is the fitted distribution's CDF evaluated at data point $x_i$
    - $F_{\\text{empirical}}(x_i) = \\frac{i}{n}$ is the empirical cumulative frequency (where $i$ is the rank of $x_i$ in sorted data)
    
    **Interpretation**:
    - **Perfect fit**: Flat line at y=0 (coinciding with x-axis) - theoretical and empirical CDFs match exactly
    - **Positive values**: Fitted distribution overestimates cumulative probability (theoretical CDF > empirical)
    - **Negative values**: Fitted distribution underestimates cumulative probability (theoretical CDF < empirical)
    - **Systematic patterns**: 
      - Curved patterns indicate shape mismatch
      - Consistent offset indicates location/scale issues
      - Oscillations suggest multimodal behavior or distribution mismatch
    
    **Advantages**:
    - Directly shows systematic biases in cumulative probability estimation
    - Easy to interpret: zero line = perfect fit
    - Complements Q-Q and P-P plots by showing signed differences
    - Useful for identifying regions where the distribution fits well or poorly
    
    #### Further Reading
    
    - **Q-Q Plots**: [Wikipedia - Q-Q Plot](https://en.wikipedia.org/wiki/Q%E2%80%93Q_plot)
    - **P-P Plots**: [Wikipedia - P-P Plot](https://en.wikipedia.org/wiki/P%E2%80%93P_plot)
    - **CDF Difference**: Related to the Kolmogorov-Smirnov test, which uses the maximum absolute difference between CDFs
    
    ---
    
    ### Combining Multiple Tests: Best Practices
    
    For reliable distribution assessment, use a multi-test approach:
    
    1. **Primary Ranking**: Use KS statistic to rank distributions (primary criterion)
    2. **Visual Inspection**: Check Q-Q, P-P, and CDF Difference plots for systematic patterns
    3. **Frequency Validation**: Use Chi-square p-value to confirm histogram matching
    4. **Tail Assessment**: Use Anderson-Darling when available (especially for tail-sensitive applications)
    5. **Model Selection**: Use AIC when comparing distributions with different parameter counts
    
    **Red flags** (consider different distribution):
    - High KS statistic (> 0.20) with low p-value
    - Systematic patterns in Q-Q, P-P, or CDF Difference plots
    - CDF Difference plot deviates significantly from y=0
    - Low Chi-square p-value (< 0.05)
    - Poor visual match in histogram + PDF overlay
    
    **Strong evidence for good fit**:
    - Low KS statistic (< 0.10) with high p-value (> 0.05)
    - Points align with diagonal in Q-Q and P-P plots
    - CDF Difference plot stays close to y=0 (x-axis)
    - High Chi-square p-value (≥ 0.05)
    - Good visual alignment in histogram + PDF overlay
    - Consistent results across multiple tests
    
    **Remember**: No single test is perfect. Use multiple tests and visual inspection for robust assessment.
    """)

with st.expander("Distribution Guide: What Each Distribution Fits", expanded=False):
    st.markdown("""
    ### Available Distributions and Their Typical Applications
    
    **1. Normal (Gaussian) Distribution**
    - **What it fits**: Symmetric, bell-shaped data with no skewness
    - **Typical use cases**: Measurement errors, heights/weights of populations, IQ scores, manufacturing tolerances
    - **Characteristics**: Mean = median = mode, defined by mean (μ) and standard deviation (σ)
    - **When to use**: When data appears symmetric and follows the central limit theorem
    
    **2. Lognormal Distribution**
    - **What it fits**: Positively skewed data where the logarithm is normally distributed
    - **Typical use cases**: Income distributions, stock prices, particle sizes, reaction times, insurance claims
    - **Characteristics**: Right-skewed, cannot take negative values, has a long tail to the right
    - **When to use**: When data is multiplicative or has multiplicative errors, often seen in biological and financial data
    
    **3. Exponential Distribution**
    - **What it fits**: Time between events in a Poisson process (constant rate)
    - **Typical use cases**: Time between failures, waiting times, inter-arrival times, radioactive decay
    - **Characteristics**: Memoryless property, constant hazard rate, right-skewed with mode at zero
    - **When to use**: For modeling waiting times or lifetimes when hazard rate is constant
    
    **4. Weibull Distribution**
    - **What it fits**: Flexible distribution for modeling lifetimes and failure rates
    - **Typical use cases**: Reliability analysis, survival analysis, material strength, wind speed data
    - **Characteristics**: Can model increasing, decreasing, or constant failure rates depending on shape parameter
    - **When to use**: When failure rate changes over time, very flexible for lifetime data
    
    **5. Gamma Distribution**
    - **What it fits**: Waiting times and positive continuous data with skewness
    - **Typical use cases**: Queueing systems, rainfall amounts, insurance losses, service times
    - **Characteristics**: Right-skewed, flexible shape depending on parameters, includes exponential as special case
    - **When to use**: For modeling sums of exponential random variables or when data is positively skewed
    
    **6. Beta Distribution**
    - **What it fits**: Probabilities, proportions, and values bounded between 0 and 1 (can be scaled)
    - **Typical use cases**: Success rates, proportions, market share, proportions of time, Bayesian analysis
    - **Characteristics**: Bounded on [0,1] (or scaled), very flexible shape (U-shaped, J-shaped, symmetric, skewed)
    - **When to use**: When data represents proportions or probabilities, or when bounded between two values
    
    **7. Triangular Distribution**
    - **What it fits**: Data with a clear mode and symmetric or asymmetric triangular shape
    - **Typical use cases**: Project management (PERT), expert opinion modeling, bounded uncertainty, simulations with limited data
    - **Characteristics**: Defined by minimum, mode, and maximum values; symmetric when mode is at center
    - **When to use**: When you have limited data but know min, most likely, and max values
    
    **8. Logistic Distribution**
    - **What it fits**: Similar to Normal but with heavier tails (more outliers)
    - **Typical use cases**: Growth models, S-curves, logit models in regression, neural networks (sigmoid), error distributions
    - **Characteristics**: Symmetric, bell-shaped like Normal but with fatter tails; S-shaped cumulative function
    - **When to use**: When data is symmetric but has more extreme values than Normal distribution allows
    
    **9. Generalized Extreme Value (GEV) Distribution**
    - **What it fits**: Extreme values, maxima or minima from large samples
    - **Typical use cases**: Climate extremes (temperature, rainfall), flood levels, financial risk (VaR), material strength, insurance claims
    - **Characteristics**: Flexible family including Gumbel, Fréchet, and Weibull types; handles heavy tails
    - **When to use**: For modeling extreme events, maximum/minimum values, or when data has very heavy tails
    
    **10. Pareto Distribution**
    - **What it fits**: Power law phenomena with many small values and few large values (80/20 rule)
    - **Typical use cases**: Income/wealth distributions, city sizes, website traffic, file sizes, scientific citations, insurance losses
    - **Characteristics**: Heavy-tailed, right-skewed, follows power law; few large values dominate
    - **When to use**: When data shows scale-free properties or follows Pareto principle (80/20 rule)
    
    **11. Uniform Distribution**
    - **What it fits**: Data with constant probability density over a bounded range (no preference for any value)
    - **Typical use cases**: Random number generation, round-robin scheduling, equal probability events, simulation models, bounded uncertainty
    - **Characteristics**: Flat PDF, symmetric, bounded on [loc, loc+scale]; equal probability for all values in range
    - **When to use**: When all values in a range are equally likely, or when modeling bounded uncertainty with no information about preference
    
    **12. Inverse Gaussian Distribution**
    - **What it fits**: Positive right-skewed data, especially lifetimes and first-passage times
    - **Typical use cases**: Lifetime data, first-passage times in Brownian motion, particle physics, reliability analysis, waiting times with drift
    - **Characteristics**: Right-skewed, always positive, has a mode to the left of mean; flexible shape
    - **When to use**: For modeling positive data with right skew, especially when dealing with first-passage times or when Normal/Gamma don't fit well
    
    **13. Burr Distribution**
    - **What it fits**: Heavy-tailed, flexible right-skewed distributions with two shape parameters
    - **Typical use cases**: Income distributions, insurance claims, survival analysis, reliability data with varying tail behavior
    - **Characteristics**: Very flexible family, can model light to very heavy tails; always positive; includes Pareto and Log-Logistic as special cases
    - **When to use**: When you need flexibility in tail behavior, especially for heavy-tailed or power-law-like data. Useful when single-parameter distributions (Pareto, Exponential) don't fit well.
    
    **14. Rayleigh Distribution**
    - **What it fits**: Positive values with right-skewed shape, especially magnitudes or distances
    - **Typical use cases**: Signal processing, wave propagation, wind speed modeling, radar cross-sections, magnitude of complex random variables
    - **Characteristics**: Always positive, right-skewed, one parameter (scale); special case of Weibull with shape=2
    - **When to use**: For modeling positive magnitudes, distances, or when data represents the magnitude of a 2D vector with normally distributed components
    
    **15. Nakagami Distribution**
    - **What it fits**: Positive values with flexible shape, especially signal amplitudes
    - **Typical use cases**: Wireless communications, fading channels, radar systems, modeling amplitude variations in signal processing
    - **Characteristics**: Always positive, flexible shape parameter (ν ≥ 0.5); includes Rayleigh and half-normal as special cases
    - **When to use**: For modeling signal amplitudes in communications, especially when you need more flexibility than Rayleigh distribution
    
    **16. Laplace (Double Exponential) Distribution**
    - **What it fits**: Symmetric data with heavier tails than normal distribution
    - **Typical use cases**: Economics (price changes), engineering (noise modeling), signal processing, Bayesian statistics, robust regression
    - **Characteristics**: Symmetric, bell-shaped but with heavier tails than Normal; defined by location and scale parameters
    - **When to use**: When data is symmetric but has more outliers than Normal distribution allows, or when you need a distribution with exponential tails
    
    **17. Gumbel (Right) Distribution**
    - **What it fits**: Extreme values, maximum values from large samples
    - **Typical use cases**: Extreme value analysis, hydrology (flood levels), meteorology (temperature extremes), material strength, insurance claims
    - **Characteristics**: Right-skewed, models maximum values; special case of GEV when shape parameter → 0
    - **When to use**: For modeling maximum values or extreme events, especially when GEV shape parameter is near zero
    
    **18. Log-Logistic Distribution**
    - **What it fits**: Positive right-skewed data, similar to lognormal but with heavier tails
    - **Typical use cases**: Survival analysis, economics (income, prices), reliability engineering, duration modeling, growth rates
    - **Characteristics**: Always positive, right-skewed, flexible shape; heavier tails than Lognormal; S-shaped hazard function
    - **When to use**: When data is positive and right-skewed but Lognormal doesn't fit well, especially in survival or reliability contexts
    
    **19. Cauchy Distribution**
    - **What it fits**: Heavy-tailed data with undefined mean and variance
    - **Typical use cases**: Physics (resonance phenomena), signal processing, finance (extreme events), robust statistics, outlier modeling
    - **Characteristics**: Symmetric, bell-shaped center but very heavy tails; undefined mean and variance; special case of Student's t with df=1
    - **When to use**: For modeling data with extreme outliers or when mean/variance are not well-defined. Useful for robust statistical methods.
    
    **20. Half-Normal Distribution**
    - **What it fits**: Positive values with normal-like shape (folded normal distribution)
    - **Typical use cases**: Reliability analysis, measurement errors (absolute values), Bayesian statistics, modeling positive deviations
    - **Characteristics**: Always positive, right-skewed; represents absolute value of normally distributed variable; one parameter (scale)
    - **When to use**: When data represents absolute values or positive deviations from a reference, especially when the underlying process is normal
    
    **21. Maxwell Distribution**
    - **What it fits**: Positive values representing speeds or magnitudes in 3D space
    - **Typical use cases**: Physics (molecular speeds in ideal gas), statistical mechanics, particle physics, 3D distance measurements
    - **Characteristics**: Always positive, right-skewed; models speed distribution of particles in 3D; one parameter (scale)
    - **When to use**: For modeling speeds or magnitudes in 3D space, especially in physics applications or when data represents 3D distances
    
    **22. Generalized Pareto Distribution**
    - **What it fits**: Extreme values, tail distributions, peaks-over-threshold data
    - **Typical use cases**: Extreme value modeling, tail risk analysis, insurance (large claims), flood analysis, financial risk (VaR), environmental extremes
    - **Characteristics**: Flexible tail behavior; can model exponential, Pareto, or bounded tails depending on shape parameter; three parameters
    - **When to use**: For modeling extreme values or tail behavior, especially in peaks-over-threshold analysis or when focusing on extreme events
    
    **23. Johnson SU Distribution**
    - **What it fits**: Very flexible distribution that can model many shapes (unbounded)
    - **Typical use cases**: Flexible modeling when other distributions don't fit, data transformation, fitting any continuous distribution shape
    - **Characteristics**: Unbounded, four parameters allowing extreme flexibility; can approximate Normal, Lognormal, and many other shapes
    - **When to use**: When you need maximum flexibility and other distributions don't fit well. Can model symmetric, skewed, or multimodal data (with appropriate parameters)
    
    **24. Johnson SB Distribution**
    - **What it fits**: Very flexible bounded distribution that can model many shapes
    - **Typical use cases**: Flexible bounded modeling, proportions with flexible shape, data on bounded intervals with complex distributions
    - **Characteristics**: Bounded on [loc, loc+scale], four parameters allowing extreme flexibility; can approximate Beta, Uniform, and many other bounded shapes
    - **When to use**: When data is bounded and you need maximum flexibility. Useful when Beta distribution is too restrictive or when bounded data has complex shapes
    
    ### Tips for Choosing a Distribution
    
    - **Check your data shape**: Is it symmetric, right-skewed, or left-skewed?
    - **Consider bounds**: Is your data bounded (e.g., 0 to 1, or always positive)?
    - **Think about the process**: What generates your data? (e.g., sums → Normal, products → Lognormal)
    - **Use automatic fitting**: Let the tool rank distributions by goodness-of-fit
    - **Visual inspection**: Look at how well the curve matches your data points on the plot
    - **KS statistic**: Lower values indicate better fit, but also consider practical interpretability
    """)

# Initialize session state
if 'data' not in st.session_state:
    st.session_state.data = None
if 'data_valid' not in st.session_state:
    st.session_state.data_valid = False


def load_test_data():
    """Load the test dataset."""
    test_path = Path(__file__).parent / "data" / "default_data.csv"
    if test_path.exists():
        df = pd.read_csv(test_path)
        return df['value'].values
    else:
        # Fallback test data
        return np.array([
            69.12, 24.77, 18.15, 6.82, 80.28, 1.76, 108.17, 3.65, 7.34, 3.78,
            1.73, 70.81, 30.46, 0.49, 1.55, 6.42, 11.15, 164, 9.57, 32.45,
            7.19, 62.88, 3.99, 709.22, 211.83, 19.53, 0.37, 9.49, 25.34, 0.09,
            64.96, 0.65, 116.66, 7.27, 11.16, 0.76, 4.93, 30.22, 80.2, 8.88,
            31.26, 151.9, 49.7, 419.89, 197.51, 25.42, 46.1, 15.48, 1.35, 262.52,
            52.86, 19.34, 38.13, 22.67, 4.64, 3.34, 21.93, 0.11, 28.37, 0.44,
            88.97, 410.42, 24.02, 11.23, 80.08, 152.55, 3.53, 18.17, 30.75,
            11.46, 56.22, 0.47, 14.35, 48.4, 14.82, 67.51, 11.68, 108.39, 8.56,
            5.15, 27.48, 351.16, 573.26, 104.2, 24.92, 1.11, 3.69, 14.36, 20.94,
            110.92, 13.19, 2.43, 6.11, 214.99, 121.87, 337.41, 252.68, 17.12,
            717.63, 48.66, 50.41, 11.73, 11.52, 10, 15.28, 1.2, 26.46, 14.65,
            43.58, 81.19, 10.86, 1778, 5.93, 22.81, 99.1, 4.26, 87.72, 9.3,
            3.26, 5.33, 17.57, 198.18, 16.69, 47.76, 5.68, 62.97, 40.84, 80.19,
            14.78, 127.89, 11.28, 14.43, 11.38, 14.77, 2.75, 453.66, 2.8, 8.82, 9.29
        ])


def parse_data_input(text_input, mode='comma'):
    """
    Parse numeric values from text input.
    
    Parameters:
    -----------
    text_input : str
        Input text to parse
    mode : str
        'comma' for comma-separated values, 'tabular' for column/tabular data
        
    Returns:
    --------
    numpy array of parsed values
    """
    values = []
    
    if mode == 'comma':
        # Split by comma and clean
        for item in text_input.split(','):
            item = item.strip()
            if item:
                try:
                    val = float(item)
                    if val >= 0:  # Only non-negative values
                        values.append(val)
                except ValueError:
                    continue
    else:  # tabular mode
        # Handle tabular data (columns from Excel, tab-separated, whitespace-separated)
        # Split by lines first
        lines = text_input.strip().split('\n')
        for line in lines:
            # Split by whitespace (handles tabs, spaces, etc.)
            items = line.split()
            for item in items:
                item = item.strip().replace(',', '')  # Remove commas that might be in numbers
                if item:
                    try:
                        val = float(item)
                        if val >= 0:  # Only non-negative values
                            values.append(val)
                    except ValueError:
                        continue
    
    return np.array(values)


def process_uploaded_file(uploaded_file):
    """
    Process uploaded file in various formats (.csv, .txt, .xlsx, .json, .parquet).
    
    Returns:
    --------
    tuple: (DataFrame, None) if successful, (None, error_message) if failed
    """
    try:
        file_extension = uploaded_file.name.lower().split('.')[-1]
        
        # Reset file pointer
        uploaded_file.seek(0)
        
        if file_extension == 'csv':
            df = pd.read_csv(uploaded_file)
        elif file_extension == 'txt':
            # Try CSV first, then fall back to plain text
            try:
                df = pd.read_csv(uploaded_file, sep=None, engine='python')
            except:
                # Read as plain text and parse
                content = uploaded_file.read().decode('utf-8')
                values = parse_data_input(content)
                df = pd.DataFrame({'Value': values})
        elif file_extension in ['xlsx', 'xls']:
            try:
                df = pd.read_excel(uploaded_file)
            except ImportError:
                return None, "Excel support requires 'openpyxl'. Install with: pip install openpyxl"
        elif file_extension == 'json':
            df = pd.read_json(uploaded_file)
        elif file_extension == 'parquet':
            try:
                df = pd.read_parquet(uploaded_file)
            except ImportError:
                return None, "Parquet support requires 'pyarrow'. Install with: pip install pyarrow"
        else:
            return None, f"Unsupported file format: .{file_extension}"
        
        return df, None
    except Exception as e:
        return None, f"Error reading file: {str(e)}"


def clean_data(data_array, remove_non_numeric=True, remove_zeros=False, clip_outliers=False):
    """
    Apply data cleaning options to a numeric array.
    
    Parameters:
    -----------
    data_array : numpy array
        Input numeric data
    remove_non_numeric : bool
        Remove NaN and infinite values
    remove_zeros : bool
        Remove zero values
    clip_outliers : bool
        Clip values above P99 percentile
        
    Returns:
    --------
    numpy array: Cleaned data
    """
    cleaned = data_array.copy()
    
    if remove_non_numeric:
        cleaned = cleaned[np.isfinite(cleaned)]
    
    if remove_zeros:
        cleaned = cleaned[cleaned != 0]
    
    if clip_outliers:
        p99 = np.percentile(cleaned, 99)
        cleaned = np.clip(cleaned, None, p99)
    
    return cleaned


def calculate_statistics(data):
    """Calculate descriptive statistics."""
    # Calculate mode
    try:
        from scipy.stats import mode as stats_mode
        mode_result = stats_mode(data, keepdims=True)
        mode_val = mode_result.mode[0] if len(mode_result.mode) > 0 else np.nan
    except:
        # Fallback: find most frequent value
        unique_vals, counts = np.unique(data, return_counts=True)
        mode_val = unique_vals[np.argmax(counts)] if len(unique_vals) > 0 else np.nan
    
    # Compute additional shape statistics
    skewness = stats.skew(data)
    kurt = stats.kurtosis(data, fisher=True)  # Fisher=True -> normal dist = 0 baseline
    
    stats_dict = {
        'Minimum': np.min(data),
        'Maximum': np.max(data),
        'Mean': np.mean(data),
        'Mode': mode_val,
        'Std Dev': np.std(data),
        'P10': np.percentile(data, 10),
        'P50 (Median)': np.percentile(data, 50),
        'P90': np.percentile(data, 90),
        'Skewness': skewness,
        'Kurtosis': kurt,
        'Count': len(data)
    }
    return stats_dict


def calculate_distribution_statistics(dist_func, params):
    """Calculate statistical parameters from a distribution."""
    try:
        # Calculate percentiles using inverse CDF (PPF)
        p10 = dist_func.ppf(0.10, *params)
        p50 = dist_func.ppf(0.50, *params)
        p90 = dist_func.ppf(0.90, *params)
        
        # Calculate mean using the distribution's mean method
        mean_val = dist_func.mean(*params)
        
        # Calculate mode - try method first, otherwise find numerically
        try:
            mode_val = dist_func.mode(*params)
            # mode() returns array for some distributions
            if isinstance(mode_val, np.ndarray):
                mode_val = mode_val[0] if len(mode_val) > 0 else np.nan
        except (AttributeError, TypeError):
            # Find mode numerically by maximizing PDF
            try:
                # Use mean and std to define search range
                mean_val_for_mode = mean_val
                std_val = dist_func.std(*params) if hasattr(dist_func, 'std') else np.std(dist_func.rvs(*params, size=1000))
                # Search around mean
                x_search = np.linspace(mean_val_for_mode - 3*std_val, mean_val_for_mode + 3*std_val, 1000)
                pdf_vals = dist_func.pdf(x_search, *params)
                mode_val = x_search[np.argmax(pdf_vals)]
            except:
                mode_val = np.nan
        
        # Calculate min and max using extreme percentiles
        try:
            min_val = dist_func.ppf(0.001, *params)  # 0.1th percentile
        except:
            try:
                min_val = dist_func.ppf(0.0, *params)
            except:
                min_val = -np.inf
        
        try:
            max_val = dist_func.ppf(0.999, *params)  # 99.9th percentile
        except:
            try:
                max_val = dist_func.ppf(1.0, *params)
            except:
                max_val = np.inf
        
        return {
            'Mean': mean_val,
            'Mode': mode_val,
            'P10': p10,
            'P50': p50,
            'P90': p90,
            'Min': min_val,
            'Max': max_val
        }
    except Exception as e:
        return None


def interpret_ks_statistic(ks_stat):
    """Provide verbal interpretation of KS statistic."""
    if ks_stat < 0.05:
        return "Excellent fit", "The fitted distribution closely matches your data."
    elif ks_stat < 0.10:
        return "Good fit", "The fitted distribution provides a good match to your data."
    elif ks_stat < 0.20:
        return "Acceptable fit", "The fitted distribution reasonably matches your data, but there may be better options."
    else:
        return "Poor fit", "The fitted distribution does not match your data well. Consider trying a different distribution."


def interpret_p_value(p_value):
    """Provide verbal interpretation of P-value."""
    if p_value > 0.10:
        return "Good fit", "Strong evidence that the data could come from this distribution."
    elif p_value > 0.05:
        return "Acceptable fit", "Moderate evidence that the data could come from this distribution."
    elif p_value > 0.01:
        return "Marginal fit", "Weak evidence - the data may not come from this distribution."
    else:
        return "Poor fit", "Strong evidence against the data coming from this distribution."


def interpret_chi2(chi2_pvalue):
    """Provide verbal interpretation of Chi-square p-value."""
    if np.isnan(chi2_pvalue):
        return "N/A", "Chi-square test not available for this distribution or sample size."
    elif chi2_pvalue >= 0.05:
        return "✅ Good fit (p ≥ 0.05)", f"The observed frequencies match expected frequencies well (p={chi2_pvalue:.4f})."
    else:
        return "❌ Poor fit (p < 0.05)", f"The observed and expected frequencies differ significantly (p={chi2_pvalue:.4f})."


def interpret_aic(aic):
    """Provide verbal interpretation of AIC."""
    if np.isnan(aic):
        return "N/A", "AIC not available for this distribution."
    else:
        return "Lower is better", "AIC allows comparing relative fit between models on the same data. Lower values indicate better fit adjusted for model complexity."


def get_distribution_formula(dist_name, params, formula_type='latex'):
    """Get distribution PDF formula in LaTeX or Excel format."""
    from utils.fitting import format_params
    
    if formula_type == 'latex':
        if dist_name == 'Normal':
            return f"$f(x) = \\frac{{1}}{{\\sigma\\sqrt{{2\\pi}}}} e^{{-\\frac{{(x-\\mu)^2}}{{2\\sigma^2}}}}$ where $\\mu={params[0]:.4f}$, $\\sigma={params[1]:.4f}$"
        elif dist_name == 'Lognormal':
            mu_log_val = np.log(params[2]) if params[2] > 0 else 0
            return f"$f(x) = \\frac{{1}}{{xs\\sqrt{{2\\pi}}}} e^{{-\\frac{{(\\ln x - \\mu)^2}}{{2s^2}}}}$ where $s={params[0]:.4f}$, $\\mu=\\ln(\\text{{scale}})={mu_log_val:.4f}$"
        elif dist_name == 'Exponential':
            lambda_val = 1/params[1] if params[1] > 0 else 0
            return f"$f(x) = \\lambda e^{{-\\lambda(x-\\text{{loc}})}}$ where $\\lambda={lambda_val:.4f}$ (rate), loc={params[0]:.4f}"
        elif dist_name == 'Gamma':
            beta_val = 1/params[2] if params[2] > 0 else 0
            return f"$f(x) = \\frac{{\\beta^{{\\alpha}}}}{{\\Gamma(\\alpha)}} x^{{\\alpha-1}} e^{{-\\beta x}}$ where $\\alpha={params[0]:.4f}$, $\\beta=1/\\text{{scale}}={beta_val:.4f}$"
        elif dist_name == 'Beta':
            return f"$f(x) = \\frac{{x^{{\\alpha-1}}(1-x)^{{\\beta-1}}}}{{B(\\alpha,\\beta)}}$ where $\\alpha={params[0]:.4f}$, $\\beta={params[1]:.4f}$"
        elif dist_name == 'Weibull (minimum)':
            return f"$f(x) = \\frac{{c}}{{\\lambda}} \\left(\\frac{{x}}{{\\lambda}}\\right)^{{c-1}} e^{{-(x/\\lambda)^c}}$ where $c={params[0]:.4f}$, $\\lambda={params[2]:.4f}$"
        elif dist_name == 'Logistic':
            return f"$f(x) = \\frac{{e^{{-(x-\\mu)/s}}}}{{s(1+e^{{-(x-\\mu)/s}})^2}}$ where $\\mu={params[0]:.4f}$, $s={params[1]:.4f}$"
        elif dist_name == 'Uniform':
            return f"$f(x) = \\frac{{1}}{{b-a}}$ for $a \\leq x \\leq b$ where $a={params[0]:.4f}$, $b={params[0]+params[1]:.4f}$"
        elif dist_name == 'Inverse Gaussian':
            return f"$f(x) = \\sqrt{{\\frac{{\\mu}}{{2\\pi x^3}}}} e^{{-\\frac{{\\mu(x-\\mu)^2}}{{2\\mu^2 x}}}}$ where $\\mu={params[0]:.4f}$"
        elif dist_name == 'Burr':
            return f"$f(x) = \\frac{{cd}}{{x}} \\left(\\frac{{x/\\beta}}{{1+(x/\\beta)^c}}\\right)^{{d+1}}$ where $c={params[0]:.4f}$, $d={params[1]:.4f}$, $\\beta={params[3]:.4f}$"
        elif dist_name == 'Rayleigh':
            return f"$f(x) = \\frac{{x}}{{s^2}} e^{{-x^2/(2s^2)}}$ where $s={params[1]:.4f}$, loc={params[0]:.4f}"
        elif dist_name == 'Nakagami':
            nu_val = params[0]
            omega_val = params[2]
            return f"$f(x) = \\frac{{2\\nu^{{\\nu}}}}{{\\Gamma(\\nu)\\Omega^{{\\nu}}}} x^{{2\\nu-1}} e^{{-\\nu x^2/\\Omega}}$ where $\\nu={nu_val:.4f}$, $\\Omega={omega_val:.4f}$, loc={params[1]:.4f}"
        elif dist_name == 'Laplace':
            return f"$f(x) = \\frac{{1}}{{2b}} e^{{-|x-\\mu|/b}}$ where $\\mu={params[0]:.4f}$, $b={params[1]:.4f}$"
        elif dist_name == 'Gumbel (Right)':
            return f"$f(x) = \\frac{{1}}{{s}} e^{{-(x-\\mu)/s}} e^{{-e^{{-(x-\\mu)/s}}}}$ where $\\mu={params[0]:.4f}$, $s={params[1]:.4f}$"
        elif dist_name == 'Log-Logistic':
            return f"$f(x) = \\frac{{(c/s)(x/s)^{{c-1}}}}{{(1+(x/s)^c)^2}}$ where $c={params[0]:.4f}$, $s={params[2]:.4f}$, loc={params[1]:.4f}"
        elif dist_name == 'Cauchy':
            return f"$f(x) = \\frac{{1}}{{\\pi s \\left[1 + \\left(\\frac{{x-\\mu}}{{s}}\\right)^2\\right]}}$ where $\\mu={params[0]:.4f}$, $s={params[1]:.4f}$"
        elif dist_name == 'Half-Normal':
            return f"$f(x) = \\frac{{\\sqrt{{2}}}}{{s\\sqrt{{\\pi}}}} e^{{-x^2/(2s^2)}}$ where $s={params[1]:.4f}$, loc={params[0]:.4f}"
        elif dist_name == 'Maxwell':
            return f"$f(x) = \\sqrt{{\\frac{{2}}{{\\pi}}}} \\frac{{x^2}}{{s^3}} e^{{-x^2/(2s^2)}}$ where $s={params[1]:.4f}$, loc={params[0]:.4f}"
        elif dist_name == 'Generalized Pareto':
            if params[0] != 0:
                return f"$f(x) = \\frac{{1}}{{s}} \\left(1 + c\\frac{{x-\\mu}}{{s}}\\right)^{{-1-1/c}}$ where $c={params[0]:.4f}$, $\\mu={params[1]:.4f}$, $s={params[2]:.4f}$"
            else:
                return f"$f(x) = \\frac{{1}}{{s}} e^{{-(x-\\mu)/s}}$ (exponential limit) where $\\mu={params[1]:.4f}$, $s={params[2]:.4f}$"
        elif dist_name == 'Johnson SU':
            return f"$f(x) = \\frac{{b}}{{s\\sqrt{{2\\pi}}}} \\frac{{1}}{{\\sqrt{{1+((x-\\mu)/s)^2}}}} e^{{-\\frac{{1}}{{2}}(a+b\\sinh^{{-1}}((x-\\mu)/s))^2}}$ where $a={params[0]:.4f}$, $b={params[1]:.4f}$, $\\mu={params[2]:.4f}$, $s={params[3]:.4f}$"
        elif dist_name == 'Johnson SB':
            return f"$f(x) = \\frac{{b}}{{s}} \\frac{{1}}{{\\sqrt{{2\\pi}}}} \\frac{{1}}{{z(1-z)}} e^{{-\\frac{{1}}{{2}}(a+b\\ln(z/(1-z)))^2}}$ where $z=(x-\\mu)/s$, $a={params[0]:.4f}$, $b={params[1]:.4f}$, $\\mu={params[2]:.4f}$, $s={params[3]:.4f}$"
        else:
            return f"PDF formula for {dist_name} (parameters: {format_params(dist_name, params)})"
    else:  # Excel
        if dist_name == 'Normal':
            return f"=NORM.DIST(x, {params[0]:.6f}, {params[1]:.6f}, FALSE)"
        elif dist_name == 'Lognormal':
            # Excel LOGNORM.DIST uses mean and std of log(x), not shape
            mu_log = np.log(params[2]) if params[2] > 0 else 0
            sigma_log = params[0] if params[0] > 0 else 1
            return f"=LOGNORM.DIST(x, {mu_log:.6f}, {sigma_log:.6f}, FALSE)"
        elif dist_name == 'Exponential':
            if params[1] > 0:
                return f"=EXPON.DIST(x-{params[0]:.6f}, {1/params[1]:.6f}, FALSE)"
            else:
                return "N/A"
        elif dist_name == 'Gamma':
            return f"=GAMMA.DIST(x, {params[0]:.6f}, {params[2]:.6f}, FALSE)"
        elif dist_name == 'Beta':
            # Excel BETA.DIST assumes scale [0,1], needs adjustment
            if params[3] > 0:
                return f"=BETA.DIST((x-{params[2]:.6f})/{params[3]:.6f}, {params[0]:.6f}, {params[1]:.6f}, FALSE, 0, 1) / {params[3]:.6f}"
            else:
                return "N/A"
        elif dist_name == 'Weibull (minimum)':
            return f"=WEIBULL.DIST(x-{params[1]:.6f}, {params[0]:.6f}, {params[2]:.6f}, FALSE)"
        elif dist_name == 'Logistic':
            # No direct Excel function, approximate with normalized formula
            return f"Excel: Use VBA or approximate with =EXP(-(x-{params[0]:.6f})/{params[1]:.6f})/({params[1]:.6f}*(1+EXP(-(x-{params[0]:.6f})/{params[1]:.6f}))^2)"
        elif dist_name == 'Uniform':
            return f"=IF(AND(x>={params[0]:.6f}, x<={params[0]+params[1]:.6f}), 1/{params[1]:.6f}, 0)"
        elif dist_name == 'Inverse Gaussian':
            # No direct Excel function
            return f"Excel: No direct function. Use VBA or approximate formula"
        elif dist_name == 'Burr':
            # No direct Excel function
            return f"Excel: No direct function. Use VBA or specialized add-ins"
        elif dist_name == 'Rayleigh':
            # No direct Excel function, but can use formula
            return f"Excel: =((x-{params[0]:.6f})/{params[1]:.6f}^2)*EXP(-((x-{params[0]:.6f})/{params[1]:.6f})^2/2)"
        elif dist_name == 'Nakagami':
            # No direct Excel function
            return f"Excel: No direct function. Use VBA or approximate with GAMMA.DIST"
        elif dist_name == 'Laplace':
            # No direct Excel function
            return f"Excel: =EXP(-ABS(x-{params[0]:.6f})/{params[1]:.6f})/(2*{params[1]:.6f})"
        elif dist_name == 'Gumbel (Right)':
            # No direct Excel function
            return f"Excel: =(1/{params[1]:.6f})*EXP(-(x-{params[0]:.6f})/{params[1]:.6f})*EXP(-EXP(-(x-{params[0]:.6f})/{params[1]:.6f}))"
        elif dist_name == 'Log-Logistic':
            # Excel uses LOGNORM.INV but not exactly same - approximate
            return f"Excel: Use VBA or approximate with =(({params[0]:.6f}/{params[2]:.6f})*((x-{params[1]:.6f})/{params[2]:.6f})^({params[0]:.6f}-1))/(1+((x-{params[1]:.6f})/{params[2]:.6f})^{params[0]:.6f})^2"
        elif dist_name == 'Cauchy':
            # No direct Excel function
            return f"Excel: =1/(PI()*{params[1]:.6f}*(1+((x-{params[0]:.6f})/{params[1]:.6f})^2))"
        elif dist_name == 'Half-Normal':
            # No direct Excel function
            return f"Excel: =SQRT(2)/({params[1]:.6f}*SQRT(PI()))*EXP(-(x-{params[0]:.6f})^2/(2*{params[1]:.6f}^2))"
        elif dist_name == 'Maxwell':
            # No direct Excel function
            return f"Excel: =SQRT(2/PI())*((x-{params[0]:.6f})^2)/({params[1]:.6f}^3)*EXP(-(x-{params[0]:.6f})^2/(2*{params[1]:.6f}^2))"
        elif dist_name == 'Generalized Pareto':
            # No direct Excel function
            if params[0] != 0:
                return f"Excel: =(1/{params[2]:.6f})*(1+{params[0]:.6f}*(x-{params[1]:.6f})/{params[2]:.6f})^(-1-1/{params[0]:.6f})"
            else:
                return f"Excel: =(1/{params[2]:.6f})*EXP(-(x-{params[1]:.6f})/{params[2]:.6f})"
        elif dist_name == 'Johnson SU':
            # No direct Excel function
            return f"Excel: No direct function. Use VBA or specialized add-ins"
        elif dist_name == 'Johnson SB':
            # No direct Excel function
            return f"Excel: No direct function. Use VBA or specialized add-ins"
        else:
            return f"Excel formula not available for {dist_name}"


def plot_qq_plot(data, dist_func, params, dist_name):
    """Create Q-Q plot (quantiles of data vs quantiles of fitted distribution)."""
    try:
        sorted_data = np.sort(data)
        n = len(sorted_data)
        
        # Theoretical quantiles from fitted distribution
        theoretical_quantiles = dist_func.ppf(np.linspace(0.01, 0.99, n), *params)
        
        # Remove any infinite or NaN values
        valid_mask = np.isfinite(theoretical_quantiles) & np.isfinite(sorted_data)
        theoretical_quantiles = theoretical_quantiles[valid_mask]
        empirical_quantiles = sorted_data[valid_mask]
        
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.scatter(theoretical_quantiles, empirical_quantiles, alpha=0.6, s=30)
        
        # Add diagonal line (perfect fit)
        min_val = min(np.min(theoretical_quantiles), np.min(empirical_quantiles))
        max_val = max(np.max(theoretical_quantiles), np.max(empirical_quantiles))
        ax.plot([min_val, max_val], [min_val, max_val], 'r--', linewidth=2, label='Perfect fit')
        
        ax.set_xlabel('Theoretical Quantiles (Fitted Distribution)', fontsize=12)
        ax.set_ylabel('Empirical Quantiles (Data)', fontsize=12)
        ax.set_title(f'Q-Q Plot: {dist_name}', fontsize=14, fontweight='bold')
        ax.legend()
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        return fig
    except Exception as e:
        return None


def plot_pp_plot(data, dist_func, params, dist_name):
    """Create P-P plot (empirical CDF vs fitted CDF)."""
    try:
        sorted_data = np.sort(data)
        n = len(sorted_data)
        
        # Empirical CDF
        empirical_cdf = np.arange(1, n + 1) / n
        
        # Theoretical CDF from fitted distribution
        theoretical_cdf = dist_func.cdf(sorted_data, *params)
        
        # Remove any infinite or NaN values
        valid_mask = np.isfinite(theoretical_cdf) & np.isfinite(empirical_cdf)
        theoretical_cdf = theoretical_cdf[valid_mask]
        empirical_cdf = empirical_cdf[valid_mask]
        
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.scatter(theoretical_cdf, empirical_cdf, alpha=0.6, s=30)
        
        # Add diagonal line (perfect fit)
        ax.plot([0, 1], [0, 1], 'r--', linewidth=2, label='Perfect fit')
        
        ax.set_xlabel('Theoretical CDF (Fitted Distribution)', fontsize=12)
        ax.set_ylabel('Empirical CDF (Data)', fontsize=12)
        ax.set_title(f'P-P Plot: {dist_name}', fontsize=14, fontweight='bold')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.legend()
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        return fig
    except Exception as e:
        return None


def plot_histogram_pdf_overlay(data, dist_func, params, dist_name, n_bins=50):
    """Create histogram with PDF overlay."""
    try:
        data_min = np.min(data)
        data_max = np.max(data)
        x_range = np.linspace(data_min, data_max, 1000)
        
        # Calculate histogram
        hist_counts, bin_edges = np.histogram(data, bins=n_bins, density=True)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
        
        # Calculate PDF
        pdf_values = dist_func.pdf(x_range, *params)
        pdf_values = pdf_values[np.isfinite(pdf_values)]
        x_range = x_range[np.isfinite(pdf_values)]
        
        fig, ax = plt.subplots(figsize=(8, 6))
        
        # Plot histogram
        ax.hist(data, bins=n_bins, density=True, alpha=0.6, color='coral', label='Empirical Data (Histogram)')
        
        # Plot PDF
        ax.plot(x_range, pdf_values, 'b-', linewidth=2, label=f'{dist_name} (Fitted PDF)')
        
        ax.set_xlabel('Value', fontsize=12)
        ax.set_ylabel('Density', fontsize=12)
        ax.set_title(f'Histogram + PDF Overlay: {dist_name}', fontsize=14, fontweight='bold')
        ax.legend()
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        return fig
    except Exception as e:
        return None


def create_excel_export(data, results, stats_dict, num_distributions=10, include_diagnostics=True):
    """
    Create a comprehensive Excel export with multiple sheets, charts, and formatted data.
    
    Parameters:
    -----------
    data : numpy array
        Input data
    results : list
        List of fitted distribution results
    stats_dict : dict
        Descriptive statistics dictionary
    num_distributions : int
        Number of top distributions to include
    include_diagnostics : bool
        Whether to include diagnostic plots
    
    Returns:
    --------
    bytes : Excel file as bytes for download
    """
    if not OPENPYXL_AVAILABLE:
        return None
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Get top N distributions
    top_results = results[:num_distributions] if len(results) >= num_distributions else results
    
    # Create sheets
    report_sheet = wb.create_sheet("Report", 0)
    data_sheet = wb.create_sheet("Data", 1)
    curves_sheet = wb.create_sheet("Fitted Curves", 2)
    stats_sheet = wb.create_sheet("Statistics", 3)
    quantiles_sheet = wb.create_sheet("Quantiles", 4)
    charts_sheet = wb.create_sheet("Charts", 5)
    
    # Define styles
    header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="1f77b4")
    section_font = Font(bold=True, size=11)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # ========== REPORT SHEET ==========
    row = 1
    
    # Header
    report_sheet.merge_cells(f'A{row}:E{row}')
    report_sheet[f'A{row}'] = "FitFitFitter - Distribution Analysis Report"
    report_sheet[f'A{row}'].font = title_font
    report_sheet[f'A{row}'].alignment = center_align
    row += 1
    
    report_sheet.merge_cells(f'A{row}:E{row}')
    report_sheet[f'A{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    report_sheet[f'A{row}'].font = Font(size=10, italic=True)
    report_sheet[f'A{row}'].alignment = center_align
    row += 2
    
    # Data Summary
    report_sheet[f'A{row}'] = "Data Summary"
    report_sheet[f'A{row}'].font = section_font
    row += 1
    
    report_sheet[f'A{row}'] = "Data Points:"
    report_sheet[f'B{row}'] = len(data)
    report_sheet[f'B{row}'].number_format = '#,##0'
    row += 1
    
    report_sheet[f'A{row}'] = "Data Range:"
    report_sheet[f'B{row}'] = f"{np.min(data):.4f} to {np.max(data):.4f}"
    row += 2
    
    # Descriptive Statistics
    report_sheet[f'A{row}'] = "Descriptive Statistics"
    report_sheet[f'A{row}'].font = section_font
    row += 1
    
    # Create descriptive stats table
    desc_stats = [
        ['Count', stats_dict['Count']],
        ['Mean', stats_dict['Mean']],
        ['Mode', stats_dict['Mode'] if not np.isnan(stats_dict['Mode']) else 'N/A'],
        ['Minimum', stats_dict['Minimum']],
        ['P10', stats_dict['P10']],
        ['P50 (Median)', stats_dict['P50 (Median)']],
        ['P90', stats_dict['P90']],
        ['Maximum', stats_dict['Maximum']],
        ['Std Dev', stats_dict['Std Dev']],
        ['Skewness', stats_dict['Skewness']],
        ['Kurtosis', stats_dict['Kurtosis']],
    ]
    
    report_sheet[f'A{row}'] = "Metric"
    report_sheet[f'B{row}'] = "Value"
    report_sheet[f'A{row}'].font = header_font
    report_sheet[f'A{row}'].fill = header_fill
    report_sheet[f'B{row}'].font = header_font
    report_sheet[f'B{row}'].fill = header_fill
    report_sheet[f'A{row}'].border = border_thin
    report_sheet[f'B{row}'].border = border_thin
    row += 1
    
    for stat_name, stat_value in desc_stats:
        report_sheet[f'A{row}'] = stat_name
        if isinstance(stat_value, (int, float)) and not np.isnan(stat_value):
            report_sheet[f'B{row}'] = stat_value
            if stat_name == 'Count':
                report_sheet[f'B{row}'].number_format = '#,##0'
            else:
                report_sheet[f'B{row}'].number_format = '0.0000'
        else:
            report_sheet[f'B{row}'] = stat_value
        report_sheet[f'A{row}'].border = border_thin
        report_sheet[f'B{row}'].border = border_thin
        if row % 2 == 0:
            report_sheet[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            report_sheet[f'B{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        row += 1
    
    row += 2
    
    # Top Fitted Distributions Table
    report_sheet[f'A{row}'] = f"Top {len(top_results)} Fitted Distributions"
    report_sheet[f'A{row}'].font = section_font
    row += 1
    
    # Table headers
    headers = ['Rank', 'Distribution', 'KS Statistic', 'P-value', 'AIC', 'Chi-square', 'Chi-square p', 'Anderson-Darling', 'Parameters', 'Excel Formula']
    for col_idx, header in enumerate(headers, 1):
        cell = report_sheet.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin
        cell.alignment = center_align
    
    row += 1
    
    # Add distribution results
    for idx, result in enumerate(top_results):
        dist_name = result['name']
        ks_stat = result.get('ks_stat', np.nan)
        p_value = result.get('p_value', np.nan)
        aic = result.get('aic', np.nan)
        chi2_stat = result.get('chi2_stat', np.nan)
        chi2_pvalue = result.get('chi2_pvalue', np.nan)
        ad_stat = result.get('ad_stat', np.nan)
        params = result.get('params', [])
        
        from utils.fitting import format_params
        params_str = format_params(dist_name, params)
        
        # Get Excel formula
        try:
            excel_formula = get_distribution_formula(dist_name, params, 'excel')
            # Check if formula is available (not "N/A" or error messages)
            if excel_formula and "not available" not in excel_formula.lower() and "no direct function" not in excel_formula.lower():
                excel_formula_str = excel_formula.replace("Excel: ", "").strip()
            else:
                excel_formula_str = "N/A"
        except:
            excel_formula_str = "N/A"
        
        values = [
            idx + 1,
            dist_name,
            ks_stat,
            p_value,
            aic,
            chi2_stat,
            chi2_pvalue,
            ad_stat,
            params_str,
            excel_formula_str
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = report_sheet.cell(row=row, column=col_idx)
            
            # Handle strings (distribution name, parameters)
            if isinstance(value, str):
                cell.value = value
            # Handle numeric values
            elif isinstance(value, (int, float)):
                if not np.isnan(value):
                    cell.value = value
                    if col_idx in [3, 4, 7]:  # KS, P-value, Chi-square p
                        cell.number_format = '0.000000'
                    elif col_idx == 5:  # AIC
                        cell.number_format = '0.00'
                    elif col_idx == 6:  # Chi-square stat
                        cell.number_format = '0.0000'
                    else:
                        cell.number_format = '0.0000'
                else:
                    cell.value = 'N/A'
            else:
                cell.value = 'N/A'
            
            cell.border = border_thin
            
            # Set alignment based on column type
            if col_idx == 10:  # Excel Formula column - left align for readability
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif col_idx > 2:  # Numeric columns - right align
                cell.alignment = right_align
            else:  # Rank and Distribution name - center align
                cell.alignment = center_align
            
            # Color code based on fit quality
            if col_idx == 3:  # KS Statistic
                if isinstance(value, (int, float)) and not np.isnan(value):
                    if value < 0.05:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
                    elif value < 0.10:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
                    elif value > 0.20:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        
        if row % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                cell = report_sheet.cell(row=row, column=col_idx)
                if not cell.fill.start_color.index:
                    cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        # Set row height for Excel formula column to allow wrapping
        report_sheet.row_dimensions[row].height = 45  # Increased height for formula wrapping
        
        row += 1
    
    # Set column widths
    report_sheet.column_dimensions['A'].width = 12
    report_sheet.column_dimensions['B'].width = 25
    report_sheet.column_dimensions['C'].width = 14
    report_sheet.column_dimensions['D'].width = 12
    report_sheet.column_dimensions['E'].width = 12
    report_sheet.column_dimensions['F'].width = 14
    report_sheet.column_dimensions['G'].width = 14
    report_sheet.column_dimensions['H'].width = 18
    report_sheet.column_dimensions['I'].width = 50
    report_sheet.column_dimensions['J'].width = 80  # Excel Formula column
    
    row += 2
    
    # Add notes about Excel formulas and charts
    report_sheet[f'A{row}'] = "Notes:"
    report_sheet[f'A{row}'].font = Font(bold=True, size=10)
    row += 1
    
    report_sheet[f'A{row}'] = "• Excel Formulas: Replace 'x' in the formula with your cell reference (e.g., A1) to calculate PDF values. Formulas marked 'N/A' are not available as direct Excel functions."
    report_sheet[f'A{row}'].font = Font(italic=True, size=9)
    report_sheet.merge_cells(f'A{row}:J{row}')
    row += 1
    
    report_sheet[f'A{row}'] = "• Charts: Interactive charts are available in the 'Fitted Curves' sheet. See cell E2 for CDF comparison and E20 for histogram + PDF overlay."
    report_sheet[f'A{row}'].font = Font(italic=True, size=9)
    report_sheet.merge_cells(f'A{row}:J{row}')
    
    # ========== DATA SHEET ==========
    data_sheet['A1'] = "Raw Data"
    data_sheet['A1'].font = section_font
    data_sheet['B1'] = "Sorted Data"
    data_sheet['B1'].font = section_font
    data_sheet['C1'] = "Empirical CDF"
    data_sheet['C1'].font = section_font
    
    sorted_data = np.sort(data)
    n = len(sorted_data)
    empirical_cdf = np.arange(1, n + 1) / n
    
    for idx in range(n):
        data_sheet.cell(row=idx+2, column=1, value=data[idx])
        data_sheet.cell(row=idx+2, column=2, value=sorted_data[idx])
        data_sheet.cell(row=idx+2, column=3, value=empirical_cdf[idx])
    
    # Add fitted CDF columns
    col = 4
    for result in top_results:
        dist_func = result['dist_func']
        params = result['params']
        dist_name = result['name']
        
        data_sheet.cell(row=1, column=col, value=f"{dist_name} CDF")
        data_sheet.cell(row=1, column=col).font = section_font
        
        try:
            fitted_cdf = dist_func.cdf(sorted_data, *params)
            for idx in range(n):
                val = fitted_cdf[idx] if np.isfinite(fitted_cdf[idx]) else np.nan
                data_sheet.cell(row=idx+2, column=col, value=val)
        except:
            pass
        
        col += 1
    
    # Format data sheet
    for col in range(1, col):
        data_sheet.cell(row=1, column=col).fill = header_fill
        data_sheet.cell(row=1, column=col).font = header_font
        data_sheet.cell(row=1, column=col).border = border_thin
    
    data_sheet.column_dimensions['A'].width = 15
    data_sheet.column_dimensions['B'].width = 15
    data_sheet.column_dimensions['C'].width = 15
    
    # ========== FITTED CURVES SHEET ==========
    # Create comprehensive x range for plotting
    x_min = np.min(data)
    x_max = np.max(data)
    x_range = np.linspace(x_min, x_max, 500)
    
    curves_sheet['A1'] = "X Values"
    curves_sheet['A1'].font = section_font
    curves_sheet['A1'].fill = header_fill
    curves_sheet['A1'].font = header_font
    
    for idx, x_val in enumerate(x_range, 2):
        curves_sheet.cell(row=idx, column=1, value=x_val)
    
    # Add empirical data (interpolate to match x_range length)
    curves_sheet['B1'] = "Empirical CDF (at X)"
    curves_sheet['B1'].font = section_font
    curves_sheet['B1'].fill = header_fill
    curves_sheet['B1'].font = header_font
    
    # Calculate empirical CDF at each x_range point
    empirical_cdf_at_x = np.searchsorted(sorted_data, x_range, side='right') / n
    
    for idx, cdf_val in enumerate(empirical_cdf_at_x, 2):
        curves_sheet.cell(row=idx, column=2, value=cdf_val)
    
    # Add fitted distribution curves
    col = 4
    for result in top_results:
        dist_func = result['dist_func']
        params = result['params']
        dist_name = result['name']
        
        # CDF values
        curves_sheet.cell(row=1, column=col, value=f"{dist_name} CDF")
        curves_sheet.cell(row=1, column=col).font = section_font
        curves_sheet.cell(row=1, column=col).fill = header_fill
        
        # PDF values
        curves_sheet.cell(row=1, column=col+1, value=f"{dist_name} PDF")
        curves_sheet.cell(row=1, column=col+1).font = section_font
        curves_sheet.cell(row=1, column=col+1).fill = header_fill
        
        try:
            # Calculate CDF and PDF for x_range
            fitted_cdf = dist_func.cdf(x_range, *params)
            fitted_pdf = dist_func.pdf(x_range, *params)
            
            for idx, (cdf_val, pdf_val) in enumerate(zip(fitted_cdf, fitted_pdf), 2):
                curves_sheet.cell(row=idx, column=col, value=cdf_val if np.isfinite(cdf_val) else None)
                curves_sheet.cell(row=idx, column=col+1, value=pdf_val if np.isfinite(pdf_val) else None)
        except:
            pass
        
        col += 2
    
    curves_sheet.column_dimensions['A'].width = 15
    curves_sheet.column_dimensions['B'].width = 15
    
    # ========== STATISTICS SHEET ==========
    stats_sheet['A1'] = "Distribution"
    stats_sheet['B1'] = "KS Statistic"
    stats_sheet['C1'] = "P-value"
    stats_sheet['D1'] = "AIC"
    stats_sheet['E1'] = "Chi-square"
    stats_sheet['F1'] = "Chi-square p"
    stats_sheet['G1'] = "Anderson-Darling"
    
    for col in range(1, 8):
        stats_sheet.cell(row=1, column=col).font = header_font
        stats_sheet.cell(row=1, column=col).fill = header_fill
        stats_sheet.cell(row=1, column=col).border = border_thin
    
    row = 2
    for result in top_results:
        stats_sheet.cell(row=row, column=1, value=result['name'])
        stats_sheet.cell(row=row, column=2, value=result.get('ks_stat', np.nan))
        stats_sheet.cell(row=row, column=3, value=result.get('p_value', np.nan))
        stats_sheet.cell(row=row, column=4, value=result.get('aic', np.nan))
        stats_sheet.cell(row=row, column=5, value=result.get('chi2_stat', np.nan))
        stats_sheet.cell(row=row, column=6, value=result.get('chi2_pvalue', np.nan))
        stats_sheet.cell(row=row, column=7, value=result.get('ad_stat', np.nan))
        
        for col in range(1, 8):
            cell = stats_sheet.cell(row=row, column=col)
            cell.border = border_thin
            if col > 1:
                cell.number_format = '0.000000'
        
        row += 1
    
    # Add parameter columns
    param_start_col = 8
    for col_idx in range(param_start_col, param_start_col + 10):
        stats_sheet.cell(row=1, column=col_idx, value=f"Param {col_idx - param_start_col + 1}")
        stats_sheet.cell(row=1, column=col_idx).font = header_font
        stats_sheet.cell(row=1, column=col_idx).fill = header_fill
    
    row = 2
    for result in top_results:
        params = result.get('params', [])
        for param_idx, param_val in enumerate(params[:10], 0):
            stats_sheet.cell(row=row, column=param_start_col + param_idx, value=param_val)
            stats_sheet.cell(row=row, column=param_start_col + param_idx).number_format = '0.000000'
        row += 1
    
    # ========== QUANTILES SHEET ==========
    quantiles_sheet['A1'] = "Distribution"
    quantile_labels = ['P10', 'P25', 'P50', 'P75', 'P90', 'Mean', 'Mode', 'Min', 'Max']
    
    for col_idx, label in enumerate(quantile_labels, 2):
        quantiles_sheet.cell(row=1, column=col_idx, value=label)
        quantiles_sheet.cell(row=1, column=col_idx).font = header_font
        quantiles_sheet.cell(row=1, column=col_idx).fill = header_fill
        quantiles_sheet.cell(row=1, column=col_idx).border = border_thin
    
    row = 2
    for result in top_results:
        dist_func = result['dist_func']
        params = result['params']
        dist_name = result['name']
        
        quantiles_sheet.cell(row=row, column=1, value=dist_name)
        quantiles_sheet.cell(row=row, column=1).border = border_thin
        
        try:
            # Calculate quantiles
            p10 = dist_func.ppf(0.10, *params)
            p25 = dist_func.ppf(0.25, *params)
            p50 = dist_func.ppf(0.50, *params)
            p75 = dist_func.ppf(0.75, *params)
            p90 = dist_func.ppf(0.90, *params)
            mean_val = dist_func.mean(*params)
            
            # Try to get mode
            try:
                mode_val = dist_func.mode(*params)
                if isinstance(mode_val, np.ndarray):
                    mode_val = mode_val[0] if len(mode_val) > 0 else np.nan
            except:
                mode_val = np.nan
            
            # Min and Max (use distribution support if available)
            try:
                min_val = dist_func.ppf(0.0001, *params) if hasattr(dist_func, 'ppf') else np.nan
                max_val = dist_func.ppf(0.9999, *params) if hasattr(dist_func, 'ppf') else np.nan
            except:
                min_val = np.nan
                max_val = np.nan
            
            values = [p10, p25, p50, p75, p90, mean_val, mode_val, min_val, max_val]
            
            for col_idx, value in enumerate(values, 2):
                cell = quantiles_sheet.cell(row=row, column=col_idx)
                if isinstance(value, (int, float)) and not np.isnan(value):
                    cell.value = value
                    cell.number_format = '0.0000'
                else:
                    cell.value = 'N/A'
                cell.border = border_thin
        except:
            for col_idx in range(2, 11):
                quantiles_sheet.cell(row=row, column=col_idx, value='N/A')
                quantiles_sheet.cell(row=row, column=col_idx).border = border_thin
        
        row += 1
    
    # Set column widths for quantiles sheet
    quantiles_sheet.column_dimensions['A'].width = 25
    for col in range(2, 11):
        quantiles_sheet.column_dimensions[get_column_letter(col)].width = 12
    
    # Freeze panes for better navigation
    report_sheet.freeze_panes = 'A2'
    data_sheet.freeze_panes = 'A2'
    curves_sheet.freeze_panes = 'A2'
    stats_sheet.freeze_panes = 'A2'
    quantiles_sheet.freeze_panes = 'A2'
    
    # ========== ADD EXCEL CHARTS ==========
    # Create a dedicated Charts sheet with properly formatted data
    chart_error_msg = None
    charts_created = False
    
    try:
        # Validate we have data to chart
        if len(x_range) < 2 or len(top_results) == 0:
            raise ValueError("Insufficient data for chart creation")
        
        # Prepare chart data in Charts sheet for better organization
        # Limit data points for charts (Excel can be slow with too many points)
        max_chart_points = min(300, len(x_range))
        
        # Calculate empirical CDF at x_range points
        sorted_data = np.sort(data)
        n = len(sorted_data)
        empirical_cdf_at_x = np.searchsorted(sorted_data, x_range[:max_chart_points], side='right') / n
        
        # Write chart data to Charts sheet
        charts_sheet['A1'] = "X Values"
        charts_sheet['B1'] = "Empirical CDF"
        charts_sheet['A1'].font = header_font
        charts_sheet['A1'].fill = header_fill
        charts_sheet['B1'].font = header_font
        charts_sheet['B1'].fill = header_fill
        
        # Write X and empirical CDF data
        for idx in range(max_chart_points):
            row = idx + 2
            x_val = x_range[idx] if idx < len(x_range) else None
            empirical_cdf_val = empirical_cdf_at_x[idx] if idx < len(empirical_cdf_at_x) else None
            
            if x_val is not None and empirical_cdf_val is not None:
                charts_sheet.cell(row=row, column=1, value=x_val)
                charts_sheet.cell(row=row, column=2, value=empirical_cdf_val)
        
        # Write fitted distribution CDF data
        chart_colors = ['4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '5B9BD5']
        num_charts = min(5, len(top_results))
        
        for idx in range(num_charts):
            result = top_results[idx]
            dist_name = result['name']
            dist_func = result['dist_func']
            params = result['params']
            
            # Column for this distribution's CDF
            cdf_col = 3 + idx
            charts_sheet.cell(row=1, column=cdf_col, value=f"{dist_name} CDF")
            charts_sheet.cell(row=1, column=cdf_col).font = header_font
            charts_sheet.cell(row=1, column=cdf_col).fill = header_fill
            
            # Calculate and write CDF values
            try:
                fitted_cdf = dist_func.cdf(x_range[:max_chart_points], *params)
                for data_idx in range(max_chart_points):
                    if data_idx < len(fitted_cdf):
                        cdf_val = fitted_cdf[data_idx] if np.isfinite(fitted_cdf[data_idx]) else None
                        if cdf_val is not None:
                            charts_sheet.cell(row=data_idx+2, column=cdf_col, value=cdf_val)
            except:
                pass
        
        # Chart 1: Cumulative Distribution Chart (on Charts sheet)
        chart1 = ScatterChart()
        chart1.title = "Cumulative Distribution Functions (CDF)"
        chart1.style = 10
        chart1.y_axis.title = 'Cumulative Probability'
        chart1.x_axis.title = 'Value'
        chart1.height = 15
        chart1.width = 20
        
        # Configure axes to show values with proper formatting
        chart1.y_axis.majorGridlines = None
        chart1.x_axis.majorGridlines = None
        
        # Set Y-axis (Cumulative Probability) to 0-1 range
        chart1.y_axis.scaling.min = 0
        chart1.y_axis.scaling.max = 1
        chart1.y_axis.majorUnit = 0.1  # Show ticks every 0.1
        
        # Auto-scale X-axis based on data
        try:
            x_min = np.min(x_range[:max_chart_points])
            x_max = np.max(x_range[:max_chart_points])
            chart1.x_axis.scaling.min = x_min
            chart1.x_axis.scaling.max = x_max
            
            # Set reasonable tick spacing for X-axis
            x_range_size = x_max - x_min
            if x_range_size > 0:
                # Try to show about 5-10 ticks
                tick_spacing = x_range_size / 8
                chart1.x_axis.majorUnit = tick_spacing
        except:
            pass
        
        # Add empirical CDF - ScatterChart needs values parameter explicitly
        x_ref = Reference(charts_sheet, min_col=1, min_row=2, max_row=max_chart_points+1)
        y_empirical = Reference(charts_sheet, min_col=2, min_row=2, max_row=max_chart_points+1)
        series_empirical = Series(values=y_empirical, xvalues=x_ref, title="Empirical CDF")
        chart1.series.append(series_empirical)
        
        # Add fitted distributions
        for idx in range(num_charts):
            result = top_results[idx]
            dist_name = result['name']
            cdf_col = 3 + idx
            
            y_cdf = Reference(charts_sheet, min_col=cdf_col, min_row=2, max_row=max_chart_points+1)
            series_cdf = Series(values=y_cdf, xvalues=x_ref, title=dist_name)
            chart1.series.append(series_cdf)
        
        # Add chart to Charts sheet
        charts_sheet.add_chart(chart1, "A3")
        charts_created = True
        
        # Chart 2: Histogram + PDF Overlay (on Charts sheet)
        # Create histogram data from actual data
        n_bins_hist = 50
        hist_counts, hist_bins = np.histogram(data, bins=n_bins_hist, density=False)
        hist_centers = (hist_bins[:-1] + hist_bins[1:]) / 2
        
        # Add histogram data to Charts sheet (start from column after distributions)
        hist_start_col = 3 + num_charts + 2  # Leave some space
        charts_sheet.cell(row=1, column=hist_start_col, value="Histogram X")
        charts_sheet.cell(row=1, column=hist_start_col+1, value="Histogram Counts")
        charts_sheet.cell(row=1, column=hist_start_col).font = header_font
        charts_sheet.cell(row=1, column=hist_start_col).fill = header_fill
        charts_sheet.cell(row=1, column=hist_start_col+1).font = header_font
        charts_sheet.cell(row=1, column=hist_start_col+1).fill = header_fill
        
        for idx, (center, count) in enumerate(zip(hist_centers, hist_counts), 2):
            charts_sheet.cell(row=idx, column=hist_start_col, value=center)
            charts_sheet.cell(row=idx, column=hist_start_col+1, value=count)
        
        # Add best fit PDF
        if len(top_results) > 0:
            best_result = top_results[0]
            best_dist_func = best_result['dist_func']
            best_params = best_result['params']
            best_name = best_result['name']
            
            charts_sheet.cell(row=1, column=hist_start_col+2, value=f"{best_name} PDF")
            charts_sheet.cell(row=1, column=hist_start_col+2).font = header_font
            charts_sheet.cell(row=1, column=hist_start_col+2).fill = header_fill
            
            try:
                pdf_values = best_dist_func.pdf(hist_centers, *best_params)
                # Scale PDF to match histogram scale (counts)
                max_pdf = np.max(pdf_values) if len(pdf_values) > 0 else 1
                max_count = np.max(hist_counts) if len(hist_counts) > 0 else 1
                if max_pdf > 0 and max_count > 0:
                    pdf_scaled = pdf_values / max_pdf * max_count
                else:
                    pdf_scaled = pdf_values
                
                for idx, pdf_val in enumerate(pdf_scaled, 2):
                    if np.isfinite(pdf_val):
                        charts_sheet.cell(row=idx, column=hist_start_col+2, value=pdf_val)
            except:
                pass
        
        # Create histogram + PDF chart
        chart2 = ScatterChart()
        chart2.title = f"Histogram and PDF Overlay - {best_name if len(top_results) > 0 else 'Best Fit'}"
        chart2.style = 10
        chart2.x_axis.title = 'Value'
        chart2.y_axis.title = 'Frequency / Density'
        chart2.height = 15
        chart2.width = 20
        
        # Configure axes to show values with proper formatting
        chart2.y_axis.majorGridlines = None
        chart2.x_axis.majorGridlines = None
        
        # Auto-scale axes based on data
        try:
            hist_x_min = np.min(hist_centers)
            hist_x_max = np.max(hist_centers)
            hist_y_max = np.max(hist_counts) * 1.1  # Add 10% padding
            
            chart2.x_axis.scaling.min = hist_x_min
            chart2.x_axis.scaling.max = hist_x_max
            chart2.y_axis.scaling.min = 0
            chart2.y_axis.scaling.max = hist_y_max
            
            # Set reasonable tick spacing
            x_range_size = hist_x_max - hist_x_min
            if x_range_size > 0:
                tick_spacing = x_range_size / 8
                chart2.x_axis.majorUnit = tick_spacing
            
            # Set Y-axis tick spacing
            if hist_y_max > 0:
                chart2.y_axis.majorUnit = hist_y_max / 8
        except:
            pass
        
        # Add histogram
        x_hist_ref = Reference(charts_sheet, min_col=hist_start_col, min_row=2, max_row=len(hist_centers)+1)
        y_hist_ref = Reference(charts_sheet, min_col=hist_start_col+1, min_row=2, max_row=len(hist_centers)+1)
        series_hist = Series(values=y_hist_ref, xvalues=x_hist_ref, title="Histogram")
        chart2.series.append(series_hist)
        
        # Add PDF curve
        if len(top_results) > 0:
            y_pdf_ref = Reference(charts_sheet, min_col=hist_start_col+2, min_row=2, max_row=len(hist_centers)+1)
            series_pdf = Series(values=y_pdf_ref, xvalues=x_hist_ref, title=f"{best_name} PDF")
            chart2.series.append(series_pdf)
        
        # Add chart to Charts sheet (below first chart)
        charts_sheet.add_chart(chart2, "A25")
        
        # Set column widths for Charts sheet
        charts_sheet.column_dimensions['A'].width = 15
        charts_sheet.column_dimensions['B'].width = 15
        for col in range(3, 3 + num_charts):
            charts_sheet.column_dimensions[get_column_letter(col)].width = 18
        
    except Exception as e:
        # Store error message for debugging
        chart_error_msg = str(e)
        charts_created = False
    
    # Add disclaimer about charts regardless of creation status
    try:
        disclaimer_row = 50
        report_sheet[f'A{disclaimer_row}'] = "IMPORTANT: Chart Visibility Disclaimer"
        report_sheet[f'A{disclaimer_row}'].font = Font(bold=True, size=11, color="FF0000")
        report_sheet.merge_cells(f'A{disclaimer_row}:J{disclaimer_row}')
        disclaimer_row += 1
        
        report_sheet[f'A{disclaimer_row}'] = "Charts have been created programmatically, but may not display correctly in all Excel versions due to compatibility issues."
        report_sheet[f'A{disclaimer_row}'].font = Font(size=10, color="FF0000")
        report_sheet.merge_cells(f'A{disclaimer_row}:J{disclaimer_row}')
        disclaimer_row += 1
        
        report_sheet[f'A{disclaimer_row}'] = "If charts are not visible:"
        report_sheet[f'A{disclaimer_row}'].font = Font(bold=True, size=10)
        report_sheet.merge_cells(f'A{disclaimer_row}:J{disclaimer_row}')
        disclaimer_row += 1
        
        report_sheet[f'A{disclaimer_row}'] = "1. All chart data is available in the 'Charts' sheet (columns A-B for empirical CDF, columns C+ for fitted distributions, histogram data in later columns)"
        report_sheet[f'A{disclaimer_row}'].font = Font(size=9)
        report_sheet.merge_cells(f'A{disclaimer_row}:J{disclaimer_row}')
        disclaimer_row += 1
        
        report_sheet[f'A{disclaimer_row}'] = "2. You can manually create charts in Excel: Select the data, go to Insert > Charts > Line Chart or Scatter Chart"
        report_sheet[f'A{disclaimer_row}'].font = Font(size=9)
        report_sheet.merge_cells(f'A{disclaimer_row}:J{disclaimer_row}')
        disclaimer_row += 1
        
        report_sheet[f'A{disclaimer_row}'] = "3. Charts are located in the 'Charts' sheet (chart at A3 and A25). All chart data is also in the 'Charts' sheet for manual chart creation."
        report_sheet[f'A{disclaimer_row}'].font = Font(size=9)
        report_sheet.merge_cells(f'A{disclaimer_row}:J{disclaimer_row}')
        disclaimer_row += 1
        
        if chart_error_msg:
            report_sheet[f'A{disclaimer_row}'] = f"4. Chart creation error: {chart_error_msg}"
            report_sheet[f'A{disclaimer_row}'].font = Font(italic=True, size=9, color="FF0000")
            report_sheet.merge_cells(f'A{disclaimer_row}:J{disclaimer_row}')
            disclaimer_row += 1
        
        if charts_created:
            report_sheet[f'A{disclaimer_row}'] = "Charts were created programmatically. If they don't appear, use the data in 'Fitted Curves' sheet to create charts manually."
            report_sheet[f'A{disclaimer_row}'].font = Font(italic=True, size=9, color="006100")
            report_sheet.merge_cells(f'A{disclaimer_row}:J{disclaimer_row}')
    except:
        pass
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def plot_cdf_difference(data, dist_func, params, dist_name):
    """Create CDF Difference plot (theoretical CDF - empirical CDF)."""
    try:
        sorted_data = np.sort(data)
        n = len(sorted_data)
        
        # Empirical CDF: cumulative frequency
        empirical_cdf = np.arange(1, n + 1) / n
        
        # Theoretical CDF from fitted distribution
        theoretical_cdf = dist_func.cdf(sorted_data, *params)
        
        # Calculate difference: theoretical - empirical
        cdf_difference = theoretical_cdf - empirical_cdf
        
        # Remove any infinite or NaN values
        valid_mask = np.isfinite(cdf_difference) & np.isfinite(sorted_data)
        cdf_difference = cdf_difference[valid_mask]
        sorted_data_valid = sorted_data[valid_mask]
        
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.plot(sorted_data_valid, cdf_difference, 'b-', linewidth=1.5, alpha=0.7, label='CDF Difference')
        ax.axhline(y=0, color='r', linestyle='--', linewidth=2, label='Perfect fit (y=0)')
        
        ax.set_xlabel('Value', fontsize=12)
        ax.set_ylabel('CDF Difference (Theoretical - Empirical)', fontsize=12)
        ax.set_title(f'CDF Difference Plot: {dist_name}', fontsize=14, fontweight='bold')
        ax.legend()
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        return fig
    except Exception as e:
        return None


def plot_multiple_distributions_plotly(data, results_list, log_scale=False, x_min=None, x_max=None, n_bins=100, num_distributions=10):
    """Plot interactive cumulative distribution with hover tooltips using Plotly."""
    # Sort data
    sorted_data = np.sort(data)
    n = len(sorted_data)
    
    # Determine x-axis limits
    if x_min is None:
        x_min = sorted_data[0]
    if x_max is None:
        x_max = sorted_data[-1]
    
    # Empirical inverse CDF: 1 - (i/n)
    empirical_y = 1 - np.arange(1, n + 1) / n
    max_cdf_val = np.max(empirical_y)
    
    # Calculate histogram
    n_bins_actual = max(n_bins, 1)
    hist_counts, hist_bins = np.histogram(data, bins=n_bins_actual, density=True, range=(x_min, x_max))
    hist_centers = (hist_bins[:-1] + hist_bins[1:]) / 2
    max_hist_val = np.max(hist_counts) if len(hist_counts) > 0 else 1
    
    # Scale histogram - initialize as empty array
    hist_scaled = np.array([])
    if max_hist_val > 0 and len(hist_counts) > 0:
        hist_scaled = hist_counts / max_hist_val * max_cdf_val * 0.5
    
    # Filter empirical data points within x-axis range
    mask = (sorted_data >= x_min) & (sorted_data <= x_max)
    plot_data = sorted_data[mask]
    plot_y = empirical_y[mask]
    
    # Create subplot with secondary y-axis
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    
    # Get seaborn color palette for selected number of distributions
    num_to_plot = min(num_distributions, len(results_list))
    sns_colors = sns.color_palette("husl", num_to_plot)
    colors = [f'#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}' 
              for r, g, b in sns_colors]
    # Plotly dash styles: 'solid', 'dot', 'dash', 'longdash', 'dashdot', 'longdashdot'
    plotly_dash_styles = ['solid', 'dash', 'dot', 'dashdot', 'longdash', 'longdashdot']
    line_styles = (plotly_dash_styles * ((num_to_plot // len(plotly_dash_styles)) + 1))[:num_to_plot]
    
    x_plot = np.linspace(x_min, x_max, 500)
    
    # Plot PDF FIRST (so it appears in the back) for best fit (rank 1)
    # Make it exactly like the manual fit plot
    if len(results_list) > 0:
        try:
            best_result = results_list[0]
            dist_func = best_result['dist_func']
            params = best_result['params']
            dist_name = best_result['name']
            rank = 1
            
            pdf_y = dist_func.pdf(x_plot, *params)
            max_pdf_val = np.max(pdf_y)
            if max_pdf_val > 0:
                pdf_scaled = pdf_y / max_pdf_val * max_cdf_val * 0.5
                
                # Convert hex color to rgba for transparent fill - use same approach as manual plot
                hex_color = colors[0].lstrip('#')
                r = int(hex_color[0:2], 16)
                g = int(hex_color[2:4], 16)
                b = int(hex_color[4:6], 16)
                fill_color_rgba = f'rgba({r}, {g}, {b}, 0.25)'  # Visible transparent fill
                line_color_rgba = f'rgba({r}, {g}, {b}, 0.4)'  # Visible transparent line
                
                # Fill area for PDF - plot FIRST so it's in the back
                fig.add_trace(
                    go.Scatter(
                        x=x_plot,
                        y=pdf_scaled,
                        mode='lines',
                        name=f'#{rank} {dist_name} (PDF, density)',
                        line=dict(color=line_color_rgba, dash='dash', width=1.5),
                        fill='tozeroy',
                        fillcolor=fill_color_rgba,
                        yaxis='y2',
                        showlegend=True,
                        legendrank=100  # Push to bottom of legend
                    ),
                    secondary_y=True
                )
        except Exception as e:
            # Don't silently fail - could help debug
            pass
    
    # Plot histogram on secondary y-axis (after PDF, but still in back)
    if max_hist_val > 0 and len(hist_centers) > 0:
        bin_width = hist_bins[1] - hist_bins[0] if len(hist_bins) > 1 else (x_max - x_min) / n_bins_actual
        fig.add_trace(
            go.Bar(
                x=hist_centers,
                y=hist_scaled,
                width=bin_width * 0.9,
                name='Empirical Histogram (density)',
                marker_color='coral',
                opacity=0.4,
                yaxis='y2',
                showlegend=True
            ),
            secondary_y=True
        )
    
    # Plot empirical data (after density functions)
    fig.add_trace(
        go.Scatter(
            x=plot_data,
            y=plot_y,
            mode='markers',
            name='Empirical Data',
            marker=dict(color='red', size=6, opacity=0.7),
            hovertemplate='<b>Value:</b> %{x:.4f}<br><b>Percentile:</b> %{y:.4%}<extra></extra>'
        ),
        secondary_y=False
    )
    
    # Plot top distributions (up to num_distributions)
    for idx, result in enumerate(results_list[:num_distributions]):
        dist_func = result['dist_func']
        params = result['params']
        dist_name = result['name']
        ks_stat = result['ks_stat']
        rank = idx + 1
        
        try:
            # Inverse CDF: 1 - CDF
            model_y = 1 - dist_func.cdf(x_plot, *params)
            
            # Calculate percentiles for hover tooltips
            percentile = (1 - model_y) * 100  # Convert to percentile
            
            # Label with rank and KS statistic
            label = f"#{rank} {dist_name} (KS={ks_stat:.4f})"
            
            # Create hover text with value and percentile
            hover_text = [f'<b>{label}</b><br>Value: {x:.4f}<br>Percentile: {p:.2f}%' 
                         for x, p in zip(x_plot, percentile)]
            
            fig.add_trace(
                go.Scatter(
                    x=x_plot,
                    y=model_y,
                    mode='lines',
                    name=label,
                    line=dict(color=colors[idx], dash=line_styles[idx], width=max(0.5, 3-idx*0.5)),
                    opacity=0.85,
                    hovertemplate='<b>%{text}</b><extra></extra>',
                    text=hover_text
                ),
                secondary_y=False
            )
            
            # Add vertical lines for best fit (rank 1)
            if rank == 1:
                try:
                    p10_val = dist_func.ppf(0.10, *params)
                    p50_val = dist_func.ppf(0.50, *params)
                    p90_val = dist_func.ppf(0.90, *params)
                    
                    # Vertical lines (not in legend)
                    fig.add_vline(x=p10_val, line_dash="dot", line_color=colors[idx], 
                                 line_width=2, opacity=0.6, annotation_text="P10",
                                 annotation_position="top", showlegend=False)
                    fig.add_vline(x=p50_val, line_dash="dash", line_color=colors[idx], 
                                 line_width=2.5, opacity=0.7, annotation_text="P50",
                                 annotation_position="top", showlegend=False)
                    fig.add_vline(x=p90_val, line_dash="dot", line_color=colors[idx], 
                                 line_width=2, opacity=0.6, annotation_text="P90",
                                 annotation_position="top", showlegend=False)
                except Exception:
                    pass
        except Exception as e:
            st.warning(f"Could not plot {dist_name}: {e}")
    
    # Update axes
    fig.update_xaxes(title_text="Value", range=[x_min, x_max])
    fig.update_yaxes(title_text="1 - Cumulative Probability", secondary_y=False)
    
    # Calculate max density for y2 axis
    max_density_val = 0
    if max_hist_val > 0 and len(hist_scaled) > 0:
        max_density_val = max(hist_scaled)
    if len(results_list) > 0 and results_list[0] is not None:
        try:
            best_dist = results_list[0]['dist_func']
            best_params = results_list[0]['params']
            x_plot_check = np.linspace(x_min, x_max, 500)
            pdf_y_check = best_dist.pdf(x_plot_check, *best_params)
            max_pdf_val_check = np.max(pdf_y_check)
            if max_pdf_val_check > 0:
                pdf_scaled_check = pdf_y_check / max_pdf_val_check * max_cdf_val * 0.5
                max_density_val = max(max_density_val, np.max(pdf_scaled_check))
        except:
            pass
    
    if max_density_val > 0:
        fig.update_yaxes(title_text="Density (scaled)", range=[0, max_density_val * 1.5], secondary_y=True)
    
    # Update layout
    fig.update_layout(
        title=f'Top {num_to_plot} Fitted Distributions - Cumulative View (Interactive)',
        hovermode='closest',
        width=1000,
        height=600,
        showlegend=True
    )
    
    if log_scale:
        fig.update_xaxis(type="log")
    
    return fig


def plot_multiple_distributions(data, results_list, log_scale=False, x_min=None, x_max=None, n_bins=100):
    """Plot inverse cumulative distribution with top N fitted distributions and density functions."""
    # Sort data
    sorted_data = np.sort(data)
    n = len(sorted_data)
    
    # Determine x-axis limits
    if x_min is None:
        x_min = sorted_data[0]
    if x_max is None:
        x_max = sorted_data[-1]
    
    # Empirical inverse CDF: 1 - (i/n)
    empirical_y = 1 - np.arange(1, n + 1) / n
    
    # Create figure with twin axes for density functions
    fig, ax = plt.subplots(figsize=(14, 8))
    ax2 = ax.twinx()
    
    # Calculate histogram for empirical data (density function)
    # Use user-specified number of bins (default 100)
    n_bins_actual = max(n_bins, 1)  # Ensure at least 1 bin
    hist_counts, hist_bins = np.histogram(data, bins=n_bins_actual, density=True, range=(x_min, x_max))
    hist_centers = (hist_bins[:-1] + hist_bins[1:]) / 2
    # Scale histogram to fit on the plot (use max of empirical inverse CDF as reference)
    max_cdf_val = np.max(empirical_y)
    max_hist_val = np.max(hist_counts) if len(hist_counts) > 0 else 1
    max_density_val = 0  # Track maximum density value for axis limit
    if max_hist_val > 0:
        hist_scaled = hist_counts / max_hist_val * max_cdf_val * 0.5
        max_density_val = max(max_density_val, np.max(hist_scaled))
        # Plot histogram bars as filled area
        bin_width = hist_bins[1] - hist_bins[0] if len(hist_bins) > 1 else (x_max - x_min) / n_bins_actual
        ax2.bar(hist_centers, hist_scaled, width=bin_width, 
               color='coral', alpha=0.4, label='Empirical Histogram (density)', zorder=0)
    
    # Filter empirical data points within x-axis range
    mask = (sorted_data >= x_min) & (sorted_data <= x_max)
    plot_data = sorted_data[mask]
    plot_y = empirical_y[mask]
    
    # Plot empirical data (inverse CDF)
    ax.scatter(plot_data, plot_y, color='red', s=25, alpha=0.7, 
               label='Empirical Data', zorder=10, edgecolors='darkred', linewidths=0.5)
    
    # Get seaborn color palette for top 5 distributions
    sns_colors = sns.color_palette("husl", 5)
    colors = sns_colors
    line_styles = ['-', '--', '-.', ':', '-']
    line_widths = [3, 2.5, 2, 1.5, 1.5]
    
    x_plot = np.linspace(x_min, x_max, 500)
    
    # Plot PDF first (so it appears in the back) for best fit (rank 1)
    if len(results_list) > 0:
        try:
            best_result = results_list[0]
            dist_func = best_result['dist_func']
            params = best_result['params']
            dist_name = best_result['name']
            
            pdf_y = dist_func.pdf(x_plot, *params)
            max_pdf_val = np.max(pdf_y)
            if max_pdf_val > 0:
                pdf_scaled = pdf_y / max_pdf_val * max_cdf_val * 0.5
                max_density_val = max(max_density_val, np.max(pdf_scaled))
                
                # Use the same color as the cumulative curve (first color from palette)
                ax2.fill_between(x_plot, 0, pdf_scaled, color=colors[0], 
                               alpha=0.15, zorder=0, label=f'#{1} {dist_name} (PDF)')  # Very transparent, at back
                ax2.plot(x_plot, pdf_scaled, color=colors[0], linestyle='--', 
                        linewidth=2, alpha=0.4, zorder=1)
        except Exception:
            pass
    
    # Plot top distributions (up to 5)
    for idx, result in enumerate(results_list[:5]):
        dist_func = result['dist_func']
        params = result['params']
        dist_name = result['name']
        ks_stat = result['ks_stat']
        rank = idx + 1
        
        try:
            # Inverse CDF: 1 - CDF
            model_y = 1 - dist_func.cdf(x_plot, *params)
            
            # Label with rank and KS statistic
            label = f"#{rank} {dist_name} (KS={ks_stat:.4f})"
            
            ax.plot(x_plot, model_y, 
                   color=colors[idx], 
                   linestyle=line_styles[idx],
                   linewidth=line_widths[idx],
                   label=label,
                   zorder=9-idx,  # Best fit on top
                   alpha=0.85)
            
            # Add percentile and mean points for the best fit (rank 1)
            if rank == 1:
                try:
                    # Calculate percentiles and mean
                    p10_val = dist_func.ppf(0.10, *params)
                    p50_val = dist_func.ppf(0.50, *params)
                    p90_val = dist_func.ppf(0.90, *params)
                    mean_val = dist_func.mean(*params)
                    
                    # Calculate y values on the cumulative curve
                    p10_y = 1 - dist_func.cdf(p10_val, *params)
                    p50_y = 1 - dist_func.cdf(p50_val, *params)
                    p90_y = 1 - dist_func.cdf(p90_val, *params)
                    mean_y = 1 - dist_func.cdf(mean_val, *params)
                    
                    # Plot points
                    ax.scatter([p10_val], [p10_y], color=colors[idx], s=100, marker='o', 
                             edgecolors='white', linewidths=1.5, zorder=15, label='P10')
                    ax.scatter([p50_val], [p50_y], color=colors[idx], s=120, marker='s', 
                             edgecolors='white', linewidths=1.5, zorder=15, label='P50 (Median)')
                    ax.scatter([p90_val], [p90_y], color=colors[idx], s=100, marker='^', 
                             edgecolors='white', linewidths=1.5, zorder=15, label='P90')
                    ax.scatter([mean_val], [mean_y], color=colors[idx], s=100, marker='D', 
                             edgecolors='white', linewidths=1.5, zorder=15, label='Mean')
                    
                    # Add vertical lines for P10, P50, P90 (not in legend)
                    ax.axvline(x=p10_val, color=colors[idx], linestyle=':', linewidth=2, 
                              alpha=0.6, zorder=14)
                    ax.axvline(x=p50_val, color=colors[idx], linestyle='--', linewidth=2.5, 
                              alpha=0.7, zorder=14)
                    ax.axvline(x=p90_val, color=colors[idx], linestyle=':', linewidth=2, 
                              alpha=0.6, zorder=14)
                    
                    # Add text labels
                    ax.annotate('P10', (p10_val, p10_y), xytext=(5, 5), 
                              textcoords='offset points', fontsize=8, color=colors[idx], 
                              fontweight='bold', zorder=16)
                    ax.annotate('P50', (p50_val, p50_y), xytext=(5, 5), 
                              textcoords='offset points', fontsize=8, color=colors[idx], 
                              fontweight='bold', zorder=16)
                    ax.annotate('P90', (p90_val, p90_y), xytext=(5, -15), 
                              textcoords='offset points', fontsize=8, color=colors[idx], 
                              fontweight='bold', zorder=16)
                    ax.annotate('Mean', (mean_val, mean_y), xytext=(5, 5), 
                              textcoords='offset points', fontsize=8, color=colors[idx], 
                              fontweight='bold', zorder=16)
                    
                except Exception as e:
                    pass  # Silently skip if percentiles can't be calculated
        except Exception as e:
            st.warning(f"Could not plot {dist_name}: {e}")
    
    # Set x-axis limits
    ax.set_xlim(x_min, x_max)
    
    # Set log scale if requested
    if log_scale:
        ax.set_xscale('log')
    
    ax.set_xlabel('Value', fontsize=13)
    ax.set_ylabel('1 - Cumulative Probability', fontsize=13)
    ax.set_title('Top 5 Fitted Distributions - Cumulative View with Density Functions', fontsize=15, fontweight='bold')
    ax.legend(loc='upper left', fontsize=10, framealpha=0.9)
    ax2.set_ylabel('Density (scaled)', fontsize=13, color='#8B7355')
    ax2.tick_params(axis='y', labelcolor='#8B7355')
    # Set density axis maximum to 50% larger than auto value
    if max_density_val > 0:
        ax2.set_ylim(0, max_density_val * 1.5)
    ax2.legend(loc='upper right', fontsize=10, framealpha=0.9)
    ax.grid(True, alpha=0.3, linestyle='--')
    
    plt.tight_layout()
    return fig


def plot_inverse_cdf(data, dist_func=None, params=None, dist_name=None, 
                     show_pdf=False, log_scale=False, x_min=None, x_max=None, n_bins=100):
    """Plot inverse cumulative distribution (1 - CDF) with density functions."""
    # Sort data
    sorted_data = np.sort(data)
    n = len(sorted_data)
    
    # Determine x-axis limits
    if x_min is None:
        x_min = sorted_data[0]
    if x_max is None:
        x_max = sorted_data[-1]
    
    # Empirical inverse CDF: 1 - (i/n)
    empirical_y = 1 - np.arange(1, n + 1) / n
    
    # Create figure with twin axes for density functions
    fig, ax1 = plt.subplots(figsize=(12, 6))
    ax2 = ax1.twinx()
    
    # Filter empirical data points within x-axis range
    mask = (sorted_data >= x_min) & (sorted_data <= x_max)
    plot_data = sorted_data[mask]
    plot_y = empirical_y[mask]
    
    # Plot empirical data (inverse CDF)
    ax1.scatter(plot_data, plot_y, color='red', s=20, alpha=0.6, 
               label='Empirical (1 - CDF)', zorder=3)
    
    # Calculate histogram for empirical data (density function)
    # Use user-specified number of bins (default 100)
    n_bins_actual = max(n_bins, 1)  # Ensure at least 1 bin
    hist_counts, hist_bins = np.histogram(data, bins=n_bins_actual, density=True, range=(x_min, x_max))
    hist_centers = (hist_bins[:-1] + hist_bins[1:]) / 2
    # Scale histogram to fit on the plot (use max of empirical inverse CDF as reference)
    max_cdf_val = np.max(empirical_y)
    max_hist_val = np.max(hist_counts) if len(hist_counts) > 0 else 1
    max_density_val = 0  # Track maximum density value for axis limit
    if max_hist_val > 0:
        hist_scaled = hist_counts / max_hist_val * max_cdf_val * 0.5  # Changed from 0.3 to 0.5
        max_density_val = max(max_density_val, np.max(hist_scaled))
        # Plot histogram bars as filled area
        bin_width = hist_bins[1] - hist_bins[0] if len(hist_bins) > 1 else (x_max - x_min) / n_bins_actual
        ax2.bar(hist_centers, hist_scaled, width=bin_width, 
               color='coral', alpha=0.4, label='Empirical Histogram (density)', zorder=0)
    
    # Plot fitted distribution if provided
    if dist_func is not None and params is not None:
        x_plot = np.linspace(x_min, x_max, 500)
        try:
            # Get seaborn color for this distribution
            sns_color = sns.color_palette("husl", 6)[0]  # First color from palette
            
            # PDF (density function) - plot first so it's in the back
            try:
                pdf_y = dist_func.pdf(x_plot, *params)
                # Normalize PDF for visualization
                max_pdf_val = np.max(pdf_y)
                if max_pdf_val > 0:
                    # Scale PDF to match histogram scaling (50% of CDF range)
                    pdf_scaled = pdf_y / max_pdf_val * max_cdf_val * 0.5
                    max_density_val = max(max_density_val, np.max(pdf_scaled))
                    # Fill area under PDF curve - very transparent, at back
                    ax2.fill_between(x_plot, 0, pdf_scaled, color=sns_color, 
                                    alpha=0.15, zorder=0, label=f'Model: {dist_name} (PDF, density)')
                    ax2.plot(x_plot, pdf_scaled, color=sns_color, linestyle='--', 
                            linewidth=2, alpha=0.4, zorder=1)
            except Exception as e:
                st.warning(f"Could not plot PDF: {e}")
            
            # Inverse CDF: 1 - CDF - plot after PDF
            model_y = 1 - dist_func.cdf(x_plot, *params)
            ax1.plot(x_plot, model_y, color=sns_color, linewidth=2, 
                    label=f'Model: {dist_name} (1 - CDF)', zorder=2)
            
            # Add percentile and mean points on the cumulative curve
            try:
                # Calculate percentiles and mean
                p10_val = dist_func.ppf(0.10, *params)
                p50_val = dist_func.ppf(0.50, *params)
                p90_val = dist_func.ppf(0.90, *params)
                mean_val = dist_func.mean(*params)
                
                # Calculate y values on the cumulative curve
                p10_y = 1 - dist_func.cdf(p10_val, *params)
                p50_y = 1 - dist_func.cdf(p50_val, *params)
                p90_y = 1 - dist_func.cdf(p90_val, *params)
                mean_y = 1 - dist_func.cdf(mean_val, *params)
                
                # Get seaborn color for markers
                sns_color_markers = sns.color_palette("husl", 6)[0]
                
                # Plot points with different markers
                ax1.scatter([p10_val], [p10_y], color=sns_color_markers, s=100, marker='o', 
                           edgecolors='white', linewidths=1.5, zorder=10, label='P10')
                ax1.scatter([p50_val], [p50_y], color=sns_color_markers, s=120, marker='s', 
                           edgecolors='white', linewidths=1.5, zorder=10, label='P50 (Median)')
                ax1.scatter([p90_val], [p90_y], color=sns_color_markers, s=100, marker='^', 
                           edgecolors='white', linewidths=1.5, zorder=10, label='P90')
                ax1.scatter([mean_val], [mean_y], color=sns_color_markers, s=100, marker='D', 
                           edgecolors='white', linewidths=1.5, zorder=10, label='Mean')
                
                # Add vertical lines for P10, P50, P90 (not in legend)
                ax1.axvline(x=p10_val, color=sns_color_markers, linestyle=':', linewidth=2, 
                           alpha=0.6, zorder=9)
                ax1.axvline(x=p50_val, color=sns_color_markers, linestyle='--', linewidth=2.5, 
                           alpha=0.7, zorder=9)
                ax1.axvline(x=p90_val, color=sns_color_markers, linestyle=':', linewidth=2, 
                           alpha=0.6, zorder=9)
                
                # Add text labels
                ax1.annotate('P10', (p10_val, p10_y), xytext=(5, 5), 
                           textcoords='offset points', fontsize=9, color=sns_color_markers, 
                           fontweight='bold', zorder=11)
                ax1.annotate('P50', (p50_val, p50_y), xytext=(5, 5), 
                           textcoords='offset points', fontsize=9, color=sns_color_markers, 
                           fontweight='bold', zorder=11)
                ax1.annotate('P90', (p90_val, p90_y), xytext=(5, -15), 
                           textcoords='offset points', fontsize=9, color=sns_color_markers, 
                           fontweight='bold', zorder=11)
                ax1.annotate('Mean', (mean_val, mean_y), xytext=(5, 5), 
                           textcoords='offset points', fontsize=9, color=sns_color_markers, 
                           fontweight='bold', zorder=11)
            except Exception as e:
                pass  # Silently skip if percentiles can't be calculated
        except Exception as e:
            st.warning(f"Could not plot distribution: {e}")
    
    # Set x-axis limits
    ax1.set_xlim(x_min, x_max)
    
    # Set log scale if requested
    if log_scale:
        ax1.set_xscale('log')
    
    ax1.set_xlabel('Value', fontsize=12)
    ax1.set_ylabel('1 - Cumulative Probability', fontsize=12)
    ax1.set_title('Inverse Cumulative Distribution Plot with Density Functions', fontsize=14, fontweight='bold')
    ax1.legend(loc='upper left')
    ax2.set_ylabel('Density (scaled)', fontsize=12, color='#8B7355')
    ax2.tick_params(axis='y', labelcolor='#8B7355')
    # Set density axis maximum to 50% larger than auto value
    if max_density_val > 0:
        ax2.set_ylim(0, max_density_val * 1.5)
    ax2.legend(loc='upper right')
    ax1.grid(True, alpha=0.3)
    
    plt.tight_layout()
    return fig


# Sidebar for data input
with st.sidebar:
    st.header("Data Input")
    
    input_method = st.radio(
        "Choose input method:",
        ["Upload File", "Paste Values", "Use Test Data"]
    )
    
    data_input = None
    
    if input_method == "Upload File":
        # Reload last data option
        if 'last_df' in st.session_state and 'last_column' in st.session_state:
            col1, col2 = st.columns([2, 1])
            with col1:
                pass  # Spacer
            with col2:
                if st.button("Reload Last Data", key="reload_last", use_container_width=True):
                    st.session_state['reload_requested'] = True
        
        uploaded_file = st.file_uploader(
            "Upload data file",
            type=['csv', 'txt', 'xlsx', 'xls', 'json', 'parquet'],
            help="Supported formats: CSV, TXT, Excel (.xlsx, .xls), JSON, Parquet"
        )
        
        # Handle reload request (when button clicked but no new file uploaded)
        if st.session_state.get('reload_requested', False) and uploaded_file is None:
            df = st.session_state.get('last_df')
            selected_col = st.session_state.get('last_column')
            
            if df is not None and selected_col is not None and selected_col in df.columns:
                raw_data = df[selected_col].values
                data_input = clean_data(raw_data, 
                                      st.session_state.get('clean_non_numeric', True),
                                      st.session_state.get('clean_zeros', False),
                                      st.session_state.get('clip_outliers', False))
                
                if data_input is not None and len(data_input) > 0:
                    st.success(f"Reloaded: {len(data_input)} values from column '{selected_col}'")
                else:
                    st.error("Unable to reload data. Please check cleaning settings.")
                    data_input = None
            else:
                st.warning("Unable to reload last data. Please upload a file.")
                data_input = None
            
            # Clear reload flag
            st.session_state['reload_requested'] = False
        
        elif uploaded_file is not None:
            # Clear reload flag when new file is uploaded
            if 'reload_requested' in st.session_state:
                st.session_state['reload_requested'] = False
            
            # Process file
            df, error_msg = process_uploaded_file(uploaded_file)
            
            if error_msg:
                st.error(f"Error: {error_msg}")
            elif df is not None:
                # Store in session state
                st.session_state['last_df'] = df
                
                # Show preview
                st.subheader("File Preview")
                st.dataframe(df.head(10), use_container_width=True)
                st.caption(f"Total rows: {len(df)}, Total columns: {len(df.columns)}")
                
                # Smart column detection
                numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
                
                if len(numeric_cols) == 0:
                    st.warning("No numeric columns found in the file. Please ensure your data contains numeric values.")
                    data_input = None
                else:
                    # Column selection
                    selected_col = st.selectbox(
                        "Select numeric column to analyze:",
                        numeric_cols,
                        index=0,
                        key="selected_column",
                        help="Choose which numeric column contains your data for distribution fitting"
                    )
                    
                    # Store selected column
                    st.session_state['last_column'] = selected_col
                    
                    # Extract selected column
                    raw_data = df[selected_col].values
                    initial_count = len(raw_data)
                    
                    # Data cleaning options
                    with st.expander("Data Import Settings", expanded=False):
                        st.markdown("**Cleaning Options:**")
                        remove_non_numeric = st.checkbox(
                            "Remove non-numeric rows",
                            value=True,
                            key="clean_non_numeric",
                            help="Remove NaN, infinite, or invalid numeric values"
                        )
                        remove_zeros = st.checkbox(
                            "Remove zero values",
                            value=False,
                            key="clean_zeros",
                            help="Useful for datasets like porosity where 0 may be invalid"
                        )
                        clip_outliers = st.checkbox(
                            "Clip outliers above P99",
                            value=False,
                            key="clip_outliers",
                            help="Remove extreme outliers by clipping values above the 99th percentile"
                        )
                    
                    # Apply cleaning (always clean, but options control what gets removed)
                    cleaned_data = clean_data(raw_data, remove_non_numeric, remove_zeros, clip_outliers)
                    final_count = len(cleaned_data)
                    
                    if final_count != initial_count:
                        st.info(f"Data cleaning: {initial_count} → {final_count} values ({initial_count - final_count} removed)")
                    
                    if final_count > 0:
                        data_input = cleaned_data
                    else:
                        st.error("All data was removed by cleaning options. Please adjust your settings.")
                        data_input = None
                    
                    if data_input is not None and len(data_input) > 0:
                        st.success(f"Ready to analyze: {len(data_input)} values from column '{selected_col}'")
                    else:
                        data_input = None
    
    elif input_method == "Paste Values":
        paste_format = st.radio(
            "Paste format:",
            ["Comma-separated", "Tabular (Column)"],
            horizontal=True,
            key="paste_format"
        )
        
        if paste_format == "Comma-separated":
            text_input = st.text_area(
                "Paste comma-separated values:",
                height=150,
                placeholder="e.g., 1.5, 2.3, 4.7, 5.1, ..."
            )
            if text_input:
                data_input = parse_data_input(text_input, mode='comma')
                if len(data_input) > 0:
                    st.success(f"Parsed {len(data_input)} values")
        else:  # Tabular
            text_input = st.text_area(
                "Paste tabular data (e.g., column from Excel):",
                height=150,
                placeholder="Paste values in a column format, one per line or separated by tabs/spaces:\n69.12\n24.77\n18.15\n..."
            )
            if text_input:
                data_input = parse_data_input(text_input, mode='tabular')
                if len(data_input) > 0:
                    st.success(f"Parsed {len(data_input)} values from tabular data")
    
    else:  # Use Test Data
        if st.button("Load Test Dataset"):
            data_input = load_test_data()
            st.success(f"Loaded {len(data_input)} test values")
    
    # Validate data
    if data_input is not None:
        # Filter non-numeric and negative values
        data_input = data_input[np.isfinite(data_input) & (data_input >= 0)]
        
        if len(data_input) < 10:
            st.error("Need at least 10 valid entries to proceed")
            st.session_state.data_valid = False
        else:
            st.session_state.data = data_input
            st.session_state.data_valid = True
            st.info(f"{len(data_input)} valid data points ready")

# Main content
if st.session_state.data_valid and st.session_state.data is not None:
    data = st.session_state.data
    
    # Descriptive Statistics
    st.header("Descriptive Statistics")
    stats_dict = calculate_statistics(data)
    
    # Custom CSS for better metric labels
    st.markdown("""
    <style>
    .big-label {
        font-size: 14px;
        font-weight: 600;
        color: #262730;
        margin-bottom: 0.25rem;
    }
    .big-value {
        font-size: 24px;
        font-weight: 700;
        color: #1f77b4;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Display statistics in requested order: count, mean, mode, min, P10, P50, P90, max
    # Using 3 columns for better balance
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(f'<div class="big-label">Count</div><div class="big-value">{stats_dict["Count"]:,}</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f'<div class="big-label">Mean</div><div class="big-value">{stats_dict["Mean"]:.4f}</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        mode_val = f"{stats_dict['Mode']:.4f}" if not np.isnan(stats_dict['Mode']) else "N/A"
        st.markdown(f'<div class="big-label">Mode</div><div class="big-value">{mode_val}</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown(f'<div class="big-label">Minimum</div><div class="big-value">{stats_dict["Minimum"]:.4f}</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f'<div class="big-label">P10</div><div class="big-value">{stats_dict["P10"]:.4f}</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f'<div class="big-label">P50 (Median)</div><div class="big-value">{stats_dict["P50 (Median)"]:.4f}</div>', unsafe_allow_html=True)
    
    with col3:
        st.markdown(f'<div class="big-label">P90</div><div class="big-value">{stats_dict["P90"]:.4f}</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f'<div class="big-label">Maximum</div><div class="big-value">{stats_dict["Maximum"]:.4f}</div>', unsafe_allow_html=True)
    
    # Display skewness, kurtosis, and std dev in Data Shape Characteristics section
    st.subheader("Data Shape Characteristics")
    col_shape1, col_shape2, col_shape3 = st.columns(3)
    with col_shape1:
        st.markdown(f'<div class="big-label">Standard Deviation</div><div class="big-value">{stats_dict["Std Dev"]:.4f}</div>', unsafe_allow_html=True)
    with col_shape2:
        st.markdown(f'<div class="big-label">Skewness</div><div class="big-value">{stats_dict["Skewness"]:.2f}</div>', unsafe_allow_html=True)
    with col_shape3:
        st.markdown(f'<div class="big-label">Kurtosis</div><div class="big-value">{stats_dict["Kurtosis"]:.2f}</div>', unsafe_allow_html=True)
    
    # Show interpretation and recommendations
    with st.expander("Interpretation of Shape Metrics", expanded=False):
        st.markdown("""
        **Skewness** and **Kurtosis** describe the *shape* of your dataset beyond mean and variance.
        
        | Metric | Interpretation | Typical Distribution |
        |:--------|:----------------|:---------------------|
        | **Skewness ≈ 0** | Symmetric distribution | Normal |
        | **Skewness > 0** | Right-skewed (long tail to the right) | Lognormal, Gamma, Weibull |
        | **Skewness < 0** | Left-skewed (long tail to the left) | Beta (left-skewed) |
        | **Kurtosis ≈ 0** | Normal-like tails | Normal |
        | **Kurtosis > 0** | Heavy tails, peaked center | Lognormal, Pareto |
        | **Kurtosis < 0** | Flat or light tails | Uniform, Triangular |
        
        **In short:**  
        - A *high positive skewness* suggests testing right-skewed distributions (e.g. Lognormal, Gamma, or Inverse Gaussian).  
        - A *high kurtosis* value (>3 if Pearson definition) implies heavier tails and potential outliers.  
        - A *negative kurtosis* indicates thinner tails (less variability at extremes), suggesting Uniform or Triangular distributions.
        
        **Heavy-tailed distributions:** Burr, Pareto, GEV, and Inverse Gaussian are useful when your data includes extreme outliers, very long tails, or rare high values.
        """)
        
        # Dynamic recommendations based on skewness
        skewness = stats_dict['Skewness']
        kurtosis = stats_dict['Kurtosis']
        if skewness > 1:
            st.info("💡 **Recommendation:** Your data is strongly right-skewed. Try Lognormal, Gamma, or Inverse Gaussian distributions.")
        elif skewness < -1:
            st.info("💡 **Recommendation:** Your data is left-skewed. Consider Beta or distributions with left-skew capability.")
        elif abs(skewness) < 0.5:
            if kurtosis < -0.5:
                st.info("💡 **Recommendation:** Your data is relatively symmetric with light tails. Normal, Logistic, or Uniform distributions may be good starting points.")
            else:
                st.info("💡 **Recommendation:** Your data is relatively symmetric. Normal or Logistic distributions may be good starting points.")
        else:
            st.info("💡 **Recommendation:** Your data shows moderate skewness. Consider a variety of distributions and compare fits.")
    
    # Automatic Distribution Fitting
    st.header("Automatic Distribution Fitting")
    
    if st.button("Fit Distributions", type="primary"):
        with st.spinner("Fitting distributions..."):
            results = rank_distributions(data)
            
            if results:
                st.session_state.fit_results = results
                st.success(f"Fitted {len(results)} distributions")
            else:
                st.error("Failed to fit any distributions")
    
    # Display ranking results
    if 'fit_results' in st.session_state:
        results = st.session_state.fit_results
        
        # Number of distributions to show selector
        max_results = len(results)
        num_distributions = st.number_input(
            "Number of distributions to show",
            min_value=1,
            max_value=max_results,
            value=min(10, max_results),
            step=1,
            key="num_distributions_to_show",
            help=f"Select how many of the top {max_results} fitted distributions to display (default: 10)"
        )
        
        # Plot with selected number of distributions
        st.subheader(f"Top {num_distributions} Fitted Distributions - Cumulative Comparison")
        
        # Plot options
        with st.expander("Plot Options", expanded=False):
            col_opt1, col_opt2, col_opt3 = st.columns(3)
            with col_opt1:
                use_auto_x = st.checkbox("Auto x-axis range", value=True, key="auto_x_best")
                if not use_auto_x:
                    data_min = float(np.min(data))
                    data_max = float(np.max(data))
                    x_min_best = st.number_input("X-axis Min", value=data_min, key="x_min_best")
                    x_max_best = st.number_input("X-axis Max", value=data_max, key="x_max_best")
                else:
                    x_min_best = None
                    x_max_best = None
            with col_opt2:
                n_bins_best = st.number_input("Number of bins", min_value=1, value=100, step=1, key="n_bins_best")
            with col_opt3:
                log_scale = st.checkbox("Log scale (x-axis)", value=False, key="best_fit_log")
        
        # Show interactive Plotly plot with hover
        fig_plotly = plot_multiple_distributions_plotly(data, results, log_scale=log_scale, 
                                                        x_min=x_min_best, x_max=x_max_best, n_bins=n_bins_best,
                                                        num_distributions=num_distributions)
        st.plotly_chart(fig_plotly, use_container_width=True)
        
        # Detailed ranking results table
        st.subheader("Detailed Results (ranked by KS statistic)")
        
        # Excel Export Button
        st.markdown("---")
        if OPENPYXL_AVAILABLE:
            # Beta notice
            st.info("**Note:** Excel export feature is currently in **BETA**. Charts may not display correctly in all Excel versions. All data is available in the exported file for manual chart creation if needed.")
            
            try:
                # Get descriptive statistics
                stats_dict = calculate_statistics(data)
                
                # Generate Excel file
                excel_bytes = create_excel_export(
                    data, 
                    results, 
                    stats_dict, 
                    num_distributions=num_distributions,
                    include_diagnostics=True
                )
                
                if excel_bytes:
                    col_exp1, col_exp2 = st.columns([1, 4])
                    with col_exp1:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"FitFitFitter_Report_{timestamp}.xlsx"
                        st.download_button(
                            label="Download Excel Report (BETA)",
                            data=excel_bytes,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Download comprehensive Excel report with all data, statistics, and fitted curves. Note: This feature is in BETA.",
                            type="primary",
                            use_container_width=True
                        )
                    with col_exp2:
                        st.caption("Export includes: Report summary, raw data, fitted curves, statistics, and quantiles. All data is formatted and ready for Excel charts. **BETA**: Charts may require manual creation in some Excel versions.")
                else:
                    st.error("Failed to generate Excel export. Please check the data and try again.")
            except Exception as e:
                st.error(f"Error generating Excel export: {str(e)}")
                st.exception(e)
        else:
            st.warning("Excel export requires openpyxl. Install with: pip install openpyxl")
        
        # Show top distributions (based on selection)
        top_n = min(num_distributions, len(results))
        
        for i, result in enumerate(results[:top_n]):
            # Get interpretations
            ks_label, ks_desc = interpret_ks_statistic(result['ks_stat'])
            p_label, p_desc = interpret_p_value(result['p_value'])
            
            with st.expander(f"#{i+1}: {result['name']} (KS = {result['ks_stat']:.4f}) - {ks_label}", 
                           expanded=(i == 0)):
                # Statistics section
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.write(f"**Rank:** {i+1}")
                    st.write(f"**KS Statistic:** {result['ks_stat']:.6f}")
                    st.caption(f"**{ks_label}**: {ks_desc}")
                    st.write(f"**P-value:** {result['p_value']:.6f}")
                    st.caption(f"**{p_label}**: {p_desc}")
                
                with col2:
                    # Display AIC with interpretation
                    aic = result.get('aic', np.nan)
                    if not np.isnan(aic):
                        aic_label, aic_desc = interpret_aic(aic)
                        st.write(f"**AIC:** {aic:.2f}")
                        st.caption(f"**{aic_label}**: {aic_desc}")
                    else:
                        st.write(f"**AIC:** N/A")
                    
                    # Display Chi-square with interpretation
                    chi2_stat = result.get('chi2_stat', np.nan)
                    chi2_pvalue = result.get('chi2_pvalue', np.nan)
                    if not np.isnan(chi2_stat) and not np.isnan(chi2_pvalue):
                        chi2_label, chi2_desc = interpret_chi2(chi2_pvalue)
                        st.write(f"**Chi-square:** {chi2_stat:.4f} (p={chi2_pvalue:.4f})")
                        st.caption(f"**{chi2_label}**: {chi2_desc}")
                    else:
                        st.write(f"**Chi-square:** N/A")
                
                with col3:
                    # Display Anderson-Darling
                    ad_stat = result.get('ad_stat', np.nan)
                    if not np.isnan(ad_stat):
                        st.write(f"**Anderson-Darling:** {ad_stat:.4f}")
                    else:
                        st.write(f"**Anderson-Darling:** N/A")
                    
                    st.write(f"**Parameters:** {format_params(result['name'], result['params'])}")
                
                # Visual Diagnostics Section
                st.markdown("---")
                st.subheader("Visual Diagnostics")
                
                st.info("""
                **Q-Q plots (Quantile-Quantile):** Good fit = points lie on diagonal line. Deviations indicate poor fit, especially in tails.
                
                **P-P plots (Probability-Probability):** Good fit = line closely follows diagonal. Shows how well cumulative probabilities match.
                
                **Histogram + PDF Overlay:** Shows how well the fitted distribution models observation frequency. Good fit = histogram bars align with PDF curve.
                
                **CDF Difference Plot:** Plots the difference between theoretical and empirical CDFs. Perfect fit = flat line at y=0 (x-axis). Deviations above/below zero indicate over/under-estimation of cumulative probabilities.
                
                Use these plots to visually inspect fit quality. Deviations from diagonal indicate poor fit in tails (Q-Q) or general shape (P-P). CDF difference plot shows systematic biases in cumulative probability estimation.
                """)
                
                # Create diagnostic plots
                dist_func = result['dist_func']
                params = result['params']
                dist_name = result['name']
                
                diag_col1, diag_col2 = st.columns(2)
                diag_col3, diag_col4 = st.columns(2)
                
                with diag_col1:
                    fig_qq = plot_qq_plot(data, dist_func, params, dist_name)
                    if fig_qq:
                        st.pyplot(fig_qq)
                    else:
                        st.info("Q-Q plot not available")
                
                with diag_col2:
                    fig_pp = plot_pp_plot(data, dist_func, params, dist_name)
                    if fig_pp:
                        st.pyplot(fig_pp)
                    else:
                        st.info("P-P plot not available")
                
                with diag_col3:
                    fig_hist = plot_histogram_pdf_overlay(data, dist_func, params, dist_name, n_bins=50)
                    if fig_hist:
                        st.pyplot(fig_hist)
                    else:
                        st.info("Histogram+PDF plot not available")
                
                with diag_col4:
                    fig_cdf_diff = plot_cdf_difference(data, dist_func, params, dist_name)
                    if fig_cdf_diff:
                        st.pyplot(fig_cdf_diff)
                    else:
                        st.info("CDF Difference plot not available")
                
                # Formula Section
                st.markdown("---")
                st.subheader("Distribution Formula")
                
                formula_col1, formula_col2 = st.columns(2)
                
                with formula_col1:
                    st.write("**LaTeX Formula:**")
                    latex_formula = get_distribution_formula(dist_name, params, 'latex')
                    st.markdown(latex_formula)
                
                with formula_col2:
                    st.write("**Excel Formula:**")
                    excel_formula = get_distribution_formula(dist_name, params, 'excel')
                    st.code(excel_formula, language=None)
                    st.caption("Replace 'x' with your variable cell reference (e.g., A1)")
                
                st.caption("""
                **How to Use in Excel:** Copy the Excel formula above, replace 'x' with your cell reference (e.g., A1), and paste into a cell to get the PDF value. 
                Note: For Lognormal, Excel uses μ and σ from log(x). For Beta, the formula includes scaling. Some distributions may require adjustments.
                """)
        
        # Separator
        st.markdown("---")
        st.markdown("")
    
    # Interactive Manual Fit Section
    st.header("Interactive Manual Fit")
    
    # Distribution selector
    dist_names = [d['name'] for d in DISTRIBUTIONS]
    selected_dist_name = st.selectbox(
        "Select Distribution:",
        dist_names,
        index=0
    )
    
    # Get distribution function
    selected_dist = next(d['func'] for d in DISTRIBUTIONS if d['name'] == selected_dist_name)
    
    # Get parameter names and bounds
    param_names = get_param_names(selected_dist_name)
    bounds = get_default_bounds(selected_dist_name, data)
    
    # Create sliders for parameters
    st.subheader("Adjust Parameters")
    params_dict = {}
    
    # Use columns for better layout
    num_params = len(param_names)
    if num_params <= 4:
        cols = st.columns(num_params)
    else:
        cols = st.columns(4)
    
    for idx, param_name in enumerate(param_names):
        col_idx = idx % 4
        with cols[col_idx]:
            if param_name in bounds:
                min_val, max_val = bounds[param_name]
                default_val = (min_val + max_val) / 2
                
                # Keys for widget synchronization
                number_input_key = f"number_{selected_dist_name}_{param_name}"
                slider_key = f"slider_{selected_dist_name}_{param_name}"
                
                # Initialize session state with default value (only if not already set)
                if number_input_key not in st.session_state:
                    st.session_state[number_input_key] = float(default_val)
                
                # Get current value from session state
                current_val = st.session_state[number_input_key]
                
                # Calculate what the slider should display (clamped to its range)
                # This is just for visual display - doesn't affect the actual parameter value
                slider_display_val = max(min_val, min(max_val, float(current_val)))
                
                # Track previous slider value to detect user interaction
                slider_prev_key = f"{slider_key}_prev"
                if slider_prev_key not in st.session_state:
                    st.session_state[slider_prev_key] = float(slider_display_val)
                
                # Slider for quick adjustment (constrained to initial bounds, secondary control)
                slider_value = st.slider(
                    f"{param_name} (slider)",
                    min_value=float(min_val),
                    max_value=float(max_val),
                    value=float(slider_display_val),
                    step=float((max_val - min_val) / 100) if (max_val - min_val) > 0 else 0.01,
                    key=slider_key,
                    help=f"Quick adjustment slider within range [{min_val:.4f}, {max_val:.4f}]. Use manual input below for values outside this range."
                )
                
                # Check if user actively moved the slider (compared to previous value)
                slider_was_moved = abs(slider_value - st.session_state[slider_prev_key]) > 1e-6
                
                # Update previous slider value for next iteration
                st.session_state[slider_prev_key] = float(slider_value)
                
                # If slider was moved, update the number input value
                if slider_was_moved:
                    st.session_state[number_input_key] = float(slider_value)
                    current_val = float(slider_value)
                
                # Number input for precise/external range values (allows any value, no min/max constraints)
                # This is the primary control - it can accept values outside the slider range
                manual_value = st.number_input(
                    f"{param_name}",
                    value=float(current_val),
                    step=float((max_val - min_val) / 200) if (max_val - min_val) > 0 else 0.01,
                    format="%.6f",
                    key=number_input_key,
                    help=f"Enter value manually. Can be outside suggested range [{min_val:.4f}, {max_val:.4f}]. This is the primary control - slider above is for quick adjustment only."
                )
                
                # Number input is always authoritative (can be outside slider range)
                # Update session state if it changed
                if abs(manual_value - current_val) > 1e-6:
                    st.session_state[number_input_key] = float(manual_value)
                
                # Use manual value for parameters (allows values outside slider range)
                params_dict[param_name] = float(manual_value)
    
    # Convert slider values to parameter tuple
    param_values = tuple(params_dict.values())
    
    # Plot inverse CDF with manual parameters
    st.subheader("Interactive Distribution Fit")
    
    # Plot options
    with st.expander("Plot Options", expanded=False):
        col_opt1, col_opt2, col_opt3 = st.columns(3)
        with col_opt1:
            use_auto_x_manual = st.checkbox("Auto x-axis range", value=True, key="auto_x_manual")
            if not use_auto_x_manual:
                data_min = float(np.min(data))
                data_max = float(np.max(data))
                x_min_manual = st.number_input("X-axis Min", value=data_min, key="x_min_manual")
                x_max_manual = st.number_input("X-axis Max", value=data_max, key="x_max_manual")
            else:
                x_min_manual = None
                x_max_manual = None
        with col_opt2:
            n_bins_manual = st.number_input("Number of bins", min_value=1, value=100, step=1, key="n_bins_manual")
        with col_opt3:
            log_scale_manual = st.checkbox("Log scale (x-axis)", value=False, key="manual_fit_log")
    
    # Create interactive Plotly plot for manual fit with hover
    try:
        sorted_data_manual = np.sort(data)
        if x_min_manual is None:
            x_min_manual = sorted_data_manual[0]
        if x_max_manual is None:
            x_max_manual = sorted_data_manual[-1]
        
        # Calculate histogram
        empirical_y_manual = 1 - np.arange(1, len(sorted_data_manual) + 1) / len(sorted_data_manual)
        max_cdf_val_manual = np.max(empirical_y_manual)
        
        n_bins_actual_manual = max(n_bins_manual, 1)
        hist_counts_manual, hist_bins_manual = np.histogram(data, bins=n_bins_actual_manual, 
                                                            density=True, range=(x_min_manual, x_max_manual))
        hist_centers_manual = (hist_bins_manual[:-1] + hist_bins_manual[1:]) / 2
        max_hist_val_manual = np.max(hist_counts_manual) if len(hist_counts_manual) > 0 else 1
        
        # Scale histogram
        if max_hist_val_manual > 0:
            hist_scaled_manual = hist_counts_manual / max_hist_val_manual * max_cdf_val_manual * 0.5
        
        x_plot_manual = np.linspace(x_min_manual, x_max_manual, 500)
        model_y_manual = 1 - selected_dist.cdf(x_plot_manual, *param_values)
        percentile_manual = (1 - model_y_manual) * 100
        
        hover_text_manual = [f'<b>{selected_dist_name}</b><br>Value: {x:.4f}<br>Percentile: {p:.2f}%' 
                            for x, p in zip(x_plot_manual, percentile_manual)]
        
        # Create subplot with secondary y-axis
        fig_plotly_manual = make_subplots(specs=[[{"secondary_y": True}]])
        
        # Get seaborn color
        sns_color_manual = sns.color_palette("husl", 6)[0]
        color_hex_manual = f'#{int(sns_color_manual[0]*255):02x}{int(sns_color_manual[1]*255):02x}{int(sns_color_manual[2]*255):02x}'
        
        # Add PDF FIRST (so it appears in the back) on secondary y-axis
        try:
            pdf_y_manual = selected_dist.pdf(x_plot_manual, *param_values)
            max_pdf_val_manual = np.max(pdf_y_manual)
            if max_pdf_val_manual > 0:
                pdf_scaled_manual = pdf_y_manual / max_pdf_val_manual * max_cdf_val_manual * 0.5
                
                # Convert hex color to rgba for transparent fill
                hex_color_manual = color_hex_manual.lstrip('#')
                r_manual = int(hex_color_manual[0:2], 16)
                g_manual = int(hex_color_manual[2:4], 16)
                b_manual = int(hex_color_manual[4:6], 16)
                fill_color_rgba_manual = f'rgba({r_manual}, {g_manual}, {b_manual}, 0.25)'  # Visible transparent fill
                line_color_rgba_manual = f'rgba({r_manual}, {g_manual}, {b_manual}, 0.4)'  # Visible transparent line
                
                fig_plotly_manual.add_trace(
                    go.Scatter(
                        x=x_plot_manual,
                        y=pdf_scaled_manual,
                        mode='lines',
                        name=f'{selected_dist_name} (PDF, density)',
                        line=dict(color=line_color_rgba_manual, dash='dash', width=1.5),
                        fill='tozeroy',
                        fillcolor=fill_color_rgba_manual,
                        yaxis='y2',
                        showlegend=True,
                        legendrank=100  # Push to bottom of legend
                    ),
                    secondary_y=True
                )
        except Exception:
            pass
        
        # Plot histogram on secondary y-axis (after PDF, but still in back)
        if max_hist_val_manual > 0 and len(hist_centers_manual) > 0:
            bin_width_manual = hist_bins_manual[1] - hist_bins_manual[0] if len(hist_bins_manual) > 1 else (x_max_manual - x_min_manual) / n_bins_actual_manual
            fig_plotly_manual.add_trace(
                go.Bar(
                    x=hist_centers_manual,
                    y=hist_scaled_manual,
                    width=bin_width_manual * 0.9,
                    name='Empirical Histogram (density)',
                    marker_color='coral',
                    opacity=0.4,
                    yaxis='y2',
                    showlegend=True
                ),
                secondary_y=True
            )
        
        # Empirical data (after density functions)
        mask_manual = (sorted_data_manual >= x_min_manual) & (sorted_data_manual <= x_max_manual)
        plot_data_manual = sorted_data_manual[mask_manual]
        plot_y_manual = empirical_y_manual[mask_manual]
        
        fig_plotly_manual.add_trace(
            go.Scatter(
                x=plot_data_manual,
                y=plot_y_manual,
                mode='markers',
                name='Empirical Data',
                marker=dict(color='red', size=6, opacity=0.7),
                hovertemplate='<b>Value:</b> %{x:.4f}<br><b>Percentile:</b> %{y:.4%}<extra></extra>'
            ),
            secondary_y=False
        )
        
        # Fitted curve - plot last (so it's in front)
        fig_plotly_manual.add_trace(
            go.Scatter(
                x=x_plot_manual,
                y=model_y_manual,
                mode='lines',
                name=f'{selected_dist_name} (1 - CDF)',
                line=dict(color=color_hex_manual, width=2),
                hovertemplate='<b>%{text}</b><extra></extra>',
                text=hover_text_manual
            ),
            secondary_y=False
        )
        
        # Add vertical lines for P10, P50, P90 (not in legend)
        try:
            p10_val_manual = selected_dist.ppf(0.10, *param_values)
            p50_val_manual = selected_dist.ppf(0.50, *param_values)
            p90_val_manual = selected_dist.ppf(0.90, *param_values)
            
            fig_plotly_manual.add_vline(x=p10_val_manual, line_dash="dot", line_color=color_hex_manual, 
                                        line_width=2, opacity=0.6, annotation_text="P10",
                                        annotation_position="top", showlegend=False)
            fig_plotly_manual.add_vline(x=p50_val_manual, line_dash="dash", line_color=color_hex_manual, 
                                        line_width=2.5, opacity=0.7, annotation_text="P50",
                                        annotation_position="top", showlegend=False)
            fig_plotly_manual.add_vline(x=p90_val_manual, line_dash="dot", line_color=color_hex_manual, 
                                        line_width=2, opacity=0.6, annotation_text="P90",
                                        annotation_position="top", showlegend=False)
        except Exception:
            pass
        
        # Calculate max density for y2 axis
        max_density_val_manual = max(hist_scaled_manual) if max_hist_val_manual > 0 and len(hist_scaled_manual) > 0 else 0
        try:
            if max_pdf_val_manual > 0:
                max_density_val_manual = max(max_density_val_manual, np.max(pdf_scaled_manual))
        except:
            pass
        
        # Update axes
        fig_plotly_manual.update_xaxes(title_text="Value", range=[x_min_manual, x_max_manual])
        fig_plotly_manual.update_yaxes(title_text="1 - Cumulative Probability", secondary_y=False)
        if max_density_val_manual > 0:
            fig_plotly_manual.update_yaxes(title_text="Density (scaled)", 
                                          range=[0, max_density_val_manual * 1.5], secondary_y=True)
        
        fig_plotly_manual.update_layout(
            title='Interactive Distribution Fit - Cumulative View',
            hovermode='closest',
            width=900,
            height=500,
            showlegend=True
        )
        
        if log_scale_manual:
            fig_plotly_manual.update_xaxis(type="log")
        
        st.plotly_chart(fig_plotly_manual, use_container_width=True)
    except Exception as e:
        st.warning(f"Could not create interactive plot: {e}")
    
    # Calculate KS statistic and distribution statistics for manual fit
    st.subheader("Fit Statistics")
    
    # Goodness of Fit metrics
    st.markdown("**Goodness of Fit:**")
    col_gof1, col_gof2, col_gof3, col_gof4 = st.columns(4)
    
    try:
        ks_stat_manual, p_value_manual = stats.kstest(
            data, 
            lambda x: selected_dist.cdf(x, *param_values)
        )
        
        # KS statistic with interpretation
        ks_label, ks_description = interpret_ks_statistic(ks_stat_manual)
        with col_gof1:
            st.markdown(f'<div class="big-label">KS Statistic</div><div class="big-value">{ks_stat_manual:.6f}</div>', unsafe_allow_html=True)
            st.caption(ks_label)
            st.caption(ks_description)
        
        # P-value with interpretation
        p_label, p_description = interpret_p_value(p_value_manual)
        with col_gof2:
            st.markdown(f'<div class="big-label">P-value</div><div class="big-value">{p_value_manual:.6f}</div>', unsafe_allow_html=True)
            st.caption(p_label)
            st.caption(p_description)
        
        # Calculate AIC for manual fit
        try:
            log_likelihood = np.sum(selected_dist.logpdf(data, *param_values))
            k = len(param_values)
            aic_manual = 2 * k - 2 * log_likelihood
            aic_label, aic_description = interpret_aic(aic_manual)
            with col_gof3:
                st.markdown(f'<div class="big-label">AIC</div><div class="big-value">{aic_manual:.2f}</div>', unsafe_allow_html=True)
                st.caption(aic_label)
                st.caption(aic_description)
        except:
            with col_gof3:
                st.markdown(f'<div class="big-label">AIC</div><div class="big-value">N/A</div>', unsafe_allow_html=True)
        
        # Placeholder for Chi-square (could be calculated if needed)
        with col_gof4:
            st.markdown(f'<div class="big-label">Chi-square</div><div class="big-value">N/A</div>', unsafe_allow_html=True)
            st.caption("Not calculated for manual fit")
        
    except Exception as e:
        st.warning(f"Could not calculate goodness-of-fit statistics: {e}")
    
    # Distribution Parameters
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**Distribution Parameters:**")
    col_param1, col_param2, col_param3 = st.columns(3)
    
    try:
        dist_stats = calculate_distribution_statistics(selected_dist, param_values)
        if dist_stats:
            with col_param1:
                st.markdown(f'<div class="big-label">Mean</div><div class="big-value">{dist_stats["Mean"]:.4f}</div>', unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                if not np.isnan(dist_stats['Mode']):
                    st.markdown(f'<div class="big-label">Mode</div><div class="big-value">{dist_stats["Mode"]:.4f}</div>', unsafe_allow_html=True)
            
            with col_param2:
                st.markdown(f'<div class="big-label">P10</div><div class="big-value">{dist_stats["P10"]:.4f}</div>', unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown(f'<div class="big-label">P50 (Median)</div><div class="big-value">{dist_stats["P50"]:.4f}</div>', unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown(f'<div class="big-label">P90</div><div class="big-value">{dist_stats["P90"]:.4f}</div>', unsafe_allow_html=True)
            
            with col_param3:
                if np.isfinite(dist_stats['Min']):
                    st.markdown(f'<div class="big-label">Min</div><div class="big-value">{dist_stats["Min"]:.4f}</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="big-label">Min</div><div class="big-value">-∞</div>', unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                if np.isfinite(dist_stats['Max']):
                    st.markdown(f'<div class="big-label">Max</div><div class="big-value">{dist_stats["Max"]:.4f}</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="big-label">Max</div><div class="big-value">+∞</div>', unsafe_allow_html=True)
        else:
            st.info("Could not calculate distribution statistics")
    except Exception as e:
        st.warning(f"Could not calculate distribution statistics: {e}")
    
else:
    st.info("Please input data using the sidebar to begin analysis.")
    
    # Show example
    with st.expander("Example Data Format"):
        st.code("""
# CSV format (one value per line):
69.12
24.77
18.15
6.82
...

# Or comma-separated:
69.12, 24.77, 18.15, 6.82, 80.28, 1.76, ...
        """)

